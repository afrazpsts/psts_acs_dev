from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
import os, re
import uuid as uuid_lib
from typing import List
from datetime import datetime, timedelta, date
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
from fastapi.responses import StreamingResponse
from utils.common_function import format_date, get_logo_url_prod
from media_manager.service import upsert_media_for_marketing


UPLOAD_FOLDER = "uploads"  

def create_marketing_service(
    db: Session,
    status_id: int = 3,
    marketing_type_id: int = None,
    announcement_type: int = None, 
    property_id: int = None,
    common_area_id: int = None,
    address: str = None,
    phone: str = None,
    country_code: str = None,
    email: str = None,
    title: str = None,
    subtext: str = None,
    description: str = None,
    duration_start_date: str = None,
    duration_end_date: str = None,
    duration_from_time: str = None,
    duration_end_time: str = None,
    location_name: str = None,
    map_link: str = None,
    website: str = None,
    terms_condition: str = None,
    start_date: str = None,
    end_date: str = None,
    cover_image: UploadFile = None
):
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        def parse_date(date_str: str):
            if not date_str:
                return None
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            raise ValueError(f"Invalid date format: {date_str}")

        today = datetime.now().date()
        parsed_start = parse_date(start_date)
        parsed_end = parse_date(end_date)

        if parsed_start and parsed_end:
            if parsed_start <= today <= parsed_end:
                status_id = 1  
            elif parsed_start > today:
                status_id = 3 
            elif parsed_end < today:
                status_id = 5 
     
        file_ext = os.path.splitext(cover_image.filename)[1]
        file_name = f"{uuid_lib.uuid4().hex}{file_ext}"
        file_path = os.path.join(UPLOAD_FOLDER, file_name)

        with open(file_path, "wb") as f:
            f.write(cover_image.file.read())

        db_file_path = f"{UPLOAD_FOLDER}/{file_name}"

        insert_sql = text("""
            INSERT INTO marketing (
                status_id, marketing_type_id, announcement_type, property_id, common_area_id, address, phone, country_code, email,
                title, subtext, description, duration_start_date, duration_end_date, duration_from_time, duration_end_time,
                location_name, map_link, website, terms_condition, cover_image, start_date, end_date, created_by, created_at, updated_at
            ) VALUES (
                :status_id, :marketing_type_id, :announcement_type, :property_id, :common_area_id, :address, :phone, :country_code, :email,
                :title, :subtext, :description, :duration_start_date, :duration_end_date, :duration_from_time, :duration_end_time,
                :location_name, :map_link, :website, :terms_condition, :cover_image, :start_date, :end_date, :created_by, NOW(), NOW()
            )
        """)

        db.execute(insert_sql, {
            "status_id": status_id,
            "marketing_type_id": marketing_type_id,
            "announcement_type": announcement_type,  
            "property_id": property_id,
            "common_area_id": common_area_id,
            "address": address,
            "phone": phone,
            "country_code": country_code,
            "email": email,
            "title": title,
            "subtext": subtext,
            "description": description,
            "duration_start_date": duration_start_date,
            "duration_end_date": duration_end_date,
            "duration_from_time": duration_from_time,
            "duration_end_time": duration_end_time,
            "location_name": location_name,
            "map_link": map_link,
            "website": website,
            "terms_condition": terms_condition,
            "cover_image": db_file_path,
            "start_date": start_date,
            "end_date": end_date,
            "created_by": "admin"
        })

        db.commit()

        last_id = db.execute(text("SELECT LAST_INSERT_ID() AS id")).scalar()
        if announcement_type is not None:
            folder = "announcement"
        else:
            folder = "marketing"
        
        try:
            upsert_media_for_marketing(
                db=db,
                marketing_id=last_id,
                file_path=db_file_path,
                folder=folder,
            )
            print(f"Successfully inserted media record for {folder} with ID: {last_id}")
        except Exception as media_error:
            print(f"Error upserting media for marketing: {media_error}")
        
        result_row = db.execute(
            text("SELECT * FROM marketing WHERE id = :id"),
            {"id": last_id}
        ).mappings().first()
        
        result_dict = dict(result_row) if result_row else None

        if result_dict:
           for key, value in result_dict.items():
            if isinstance(value, datetime):
             result_dict[key] = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, timedelta):
             result_dict[key] = str(value)

        return {
            "status": 200,
            "message": "Marketing created successfully.",
            "data": result_dict  
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



def update_marketing_service(
    db: Session,
    marketing_id: int,
    status_id: int = None,
    marketing_type_id: int = None,
    announcement_type: int = None,
    property_id: int = None,
    common_area_id: int = None,
    address: str = None,
    phone: str = None,
    country_code: str = None,
    email: str = None,
    title: str = None,
    subtext: str = None,
    description: str = None,
    duration_start_date: str = None,
    duration_end_date: str = None,
    duration_from_time: str = None,
    duration_end_time: str = None,
    location_name: str = None,
    map_link: str = None,
    website: str = None,
    terms_condition: str = None,
    start_date: str = None,
    end_date: str = None,
    cover_image: UploadFile = None
):
    try:
        marketing = db.execute(
            text("SELECT * FROM marketing WHERE id = :id"),
            {"id": marketing_id}
        ).mappings().first()

        if not marketing:
            raise HTTPException(status_code=404, detail="Marketing record not found")

        def keep_or_update(new_val, old_val):
            return new_val if new_val not in [None, ""] else old_val

        def parse_date(date_str: str):
            if not date_str:
                return None
            if isinstance(date_str, (datetime, date)):
                return date_str.date() if isinstance(date_str, datetime) else date_str
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            raise ValueError(f"Invalid date format: {date_str}")

        parsed_start = parse_date(keep_or_update(start_date, marketing["start_date"]))
        parsed_end = parse_date(keep_or_update(end_date, marketing["end_date"]))
        today = datetime.now().date()

        if parsed_start and parsed_end:
            if parsed_start <= today <= parsed_end:
                status_id = 1   
            elif parsed_start > today:
                status_id = 3   
            elif parsed_end < today:
                status_id = 5   
        else:
            status_id = keep_or_update(status_id, marketing["status_id"])

        db_file_path = marketing["cover_image"]
        file_path = None
        media_updated = False
        
        if cover_image:
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            file_ext = os.path.splitext(cover_image.filename)[1]
            file_name = f"{uuid_lib.uuid4().hex}{file_ext}"
            file_path = os.path.join(UPLOAD_FOLDER, file_name)
            with open(file_path, "wb") as f:
                f.write(cover_image.file.read())
            db_file_path = f"{UPLOAD_FOLDER}/{file_name}"
            media_updated = True
            
            
            if marketing["cover_image"] and os.path.exists(marketing["cover_image"]):
                try:
                    os.remove(marketing["cover_image"])
                    print(f"Deleted old cover image: {marketing['cover_image']}")
                except Exception as delete_error:
                    print(f"Error deleting old cover image: {delete_error}")

        update_sql = text("""
            UPDATE marketing SET
                status_id = :status_id,
                marketing_type_id = :marketing_type_id,
                announcement_type = :announcement_type,
                property_id = :property_id,
                common_area_id = :common_area_id,
                address = :address,
                phone = :phone,
                country_code = :country_code,
                email = :email,
                title = :title,
                subtext = :subtext,
                description = :description,
                duration_start_date = :duration_start_date,
                duration_end_date = :duration_end_date,
                duration_from_time = :duration_from_time,
                duration_end_time = :duration_end_time,
                location_name = :location_name,
                map_link = :map_link,
                website = :website,
                terms_condition = :terms_condition,
                cover_image = :cover_image,
                start_date = :start_date,
                end_date = :end_date,
                updated_at = NOW()
            WHERE id = :id
        """)

        db.execute(update_sql, {
            "id": marketing_id,
            "status_id": status_id,
            "marketing_type_id": keep_or_update(marketing_type_id, marketing["marketing_type_id"]),
            "announcement_type": keep_or_update(announcement_type, marketing.get("announcement_type")),
            "property_id": keep_or_update(property_id, marketing["property_id"]),
            "common_area_id": keep_or_update(common_area_id, marketing["common_area_id"]),
            "address": keep_or_update(address, marketing["address"]),
            "phone": keep_or_update(phone, marketing["phone"]),
            "country_code": keep_or_update(country_code, marketing["country_code"]),
            "email": keep_or_update(email, marketing["email"]),
            "title": keep_or_update(title, marketing["title"]),
            "subtext": keep_or_update(subtext, marketing["subtext"]),
            "description": keep_or_update(description, marketing["description"]),
            "duration_start_date": keep_or_update(duration_start_date, marketing["duration_start_date"]),
            "duration_end_date": keep_or_update(duration_end_date, marketing["duration_end_date"]),
            "duration_from_time": keep_or_update(duration_from_time, marketing["duration_from_time"]),
            "duration_end_time": keep_or_update(duration_end_time, marketing["duration_end_time"]),
            "location_name": keep_or_update(location_name, marketing["location_name"]),
            "map_link": keep_or_update(map_link, marketing["map_link"]),
            "website": keep_or_update(website, marketing["website"]),
            "terms_condition": keep_or_update(terms_condition, marketing["terms_condition"]),
            "cover_image": db_file_path,
            "start_date": keep_or_update(start_date, marketing["start_date"]),
            "end_date": keep_or_update(end_date, marketing["end_date"])
        })

        db.commit()

        if media_updated:
            
            final_announcement_type = keep_or_update(announcement_type, marketing.get("announcement_type"))
            folder = "announcement" if final_announcement_type is not None else "marketing"
            
            try:
                upsert_media_for_marketing(
                    db=db,
                    marketing_id=marketing_id,
                    file_path=db_file_path,
                    folder=folder,
                )
                print(f"Successfully updated media record for {folder} with ID: {marketing_id}")
            except Exception as media_error:
                print(f"Error upserting media for marketing: {media_error}")
            

        
        updated_result = db.execute(text("""
            SELECT 
                id,
                status_id,
                marketing_type_id,
                announcement_type,
                property_id,
                common_area_id,
                address,
                phone,
                country_code,
                email,
                title,
                subtext,
                description,
                duration_start_date,
                duration_end_date,
                duration_from_time,
                duration_end_time,
                location_name,
                map_link,
                website,
                terms_condition,
                cover_image,
                start_date,
                end_date,
                created_by,
                created_at,
                updated_at
            FROM marketing
            WHERE id = :id
        """), {"id": marketing_id}).mappings().first()
        
        updated_dict = {}
        if updated_result:
            for key, value in updated_result.items():
                if isinstance(value, datetime):
                    updated_dict[key] = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):
                    updated_dict[key] = value.strftime("%Y-%m-%d")
                elif isinstance(value, timedelta):
                    total_seconds = int(value.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    updated_dict[key] = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    updated_dict[key] = value

        return {
            "status": 200,
            "message": "Marketing updated successfully.",
            "data": updated_dict
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def format_marketing_record(record: dict) -> dict:
    """
    Format marketing record with image object similar to announcements
    """
    image_obj = None
    if record.get("image_id"):
        image_obj = {
            "id": record.get("image_id"),
            "folder": record.get("image_folder"),
            "name": record.get("image_name"),
            "path": record.get("image_path"),
            "module_name": record.get("image_module_name"),
            "reference_id": record.get("image_reference_id"),
            "file_type": record.get("image_file_type"),
            "file_size": record.get("image_file_size"),
            "created_at": record.get("image_created_at").strftime("%Y-%m-%d %H:%M:%S") if record.get("image_created_at") else None,
            "updated_at": record.get("image_updated_at").strftime("%Y-%m-%d %H:%M:%S") if record.get("image_updated_at") else None
        }
    
    formatted_record = {
        "id": record.get("id"),
        "status_id": record.get("status_id"),
        "marketing_type_id": record.get("marketing_type_id"),
        "announcement_type": record.get("announcement_type"),
        "property_id": record.get("property_id"),
        "common_area_id": record.get("common_area_id"),
        "address": record.get("address"),
        "phone": record.get("phone"),
        "country_code": record.get("country_code"),
        "email": record.get("email"),
        "title": record.get("title"),
        "subtext": record.get("subtext"),
        "description": record.get("description"),
        "duration_start_date": record.get("duration_start_date"),
        "duration_end_date": record.get("duration_end_date"),
        "duration_from_time": str(record.get("duration_from_time")) if record.get("duration_from_time") else None,
        "duration_end_time": str(record.get("duration_end_time")) if record.get("duration_end_time") else None,
        "location_name": record.get("location_name"),
        "map_link": record.get("map_link"),
        "website": record.get("website"),
        "terms_condition": record.get("terms_condition"),
        "cover_image": record.get("cover_image"),
        "start_date": record.get("start_date"),
        "end_date": record.get("end_date"),
        "created_by": record.get("created_by"),
        "created_at": record.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if record.get("created_at") else None,
        "updated_at": record.get("updated_at").strftime("%Y-%m-%d %H:%M:%S") if record.get("updated_at") else None,
        "status_name": record.get("status_name"),
        "marketing_type_name": record.get("marketing_type_name"),
        "image": image_obj 
    }
    
    
    return formatted_record

def format_time_value(time_value):
    """Format time value to HH:MM:SS string"""
    if time_value is None:
        return None
    
    if isinstance(time_value, (int, float)):
        hours = int(time_value // 3600)
        minutes = int((time_value % 3600) // 60)
        seconds = int(time_value % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    return str(time_value)

def format_announcement_record(record):
    """Format marketing record to exclude null fields and add announcement_name"""
    announcement_type_map = {
        1: "General",
        2: "Building Alert"
    }
    
    formatted_record = {}
    
    for key, value in record.items():
        if value is not None:
            formatted_record[key] = value
    
    if record.get('duration_from_time') is not None:
        formatted_record['duration_from_time'] = format_time_value(record['duration_from_time'])
    
    if record.get('duration_end_time') is not None:
        formatted_record['duration_end_time'] = format_time_value(record['duration_end_time'])
    
    if record.get('start_date') and hasattr(record['start_date'], 'strftime'):
        formatted_record['start_date'] = record['start_date'].strftime("%Y-%m-%dT%H:%M:%SZ")
    
    if record.get('end_date') and hasattr(record['end_date'], 'strftime'):
        formatted_record['end_date'] = record['end_date'].strftime("%Y-%m-%dT%H:%M:%SZ")
    
    if record.get('announcement_type') is not None:
        formatted_record['announcement_name'] = announcement_type_map.get(
            record['announcement_type'], 
            "Unknown"
        )

    image_id = record.get('image_id')
    if image_id is not None:
        image_obj = {
            "id": image_id,
            "folder": record.get("image_folder"),
            "name": record.get("image_name"),
            "path": record.get("image_path"),
            "module_name": record.get("image_module_name"),
            "reference_id": record.get("image_reference_id"),
            "file_type": record.get("image_file_type"),
            "file_size": record.get("image_file_size"),
            "created_at": record.get("image_created_at"),
            "updated_at": record.get("image_updated_at"),
        }
        formatted_record["image"] = image_obj
    else:
        formatted_record["image"] = None

    for k in list(formatted_record.keys()):
        if k.startswith("image_"):
            formatted_record.pop(k, None)

    return formatted_record

def list_marketing_service(
    db: Session,
    id: int = None,
    status_id: int = None,
    marketing_type_id: int = None,
    start_date: datetime = None,
    end_date: datetime = None,
    page: int = 1,
    record_count: int = 10
):
    try:
        base_query = """
            SELECT 
                m.*,
                ms.name AS status_name,
                mt.key AS marketing_type_name,
                mm.id AS image_id,
                mm.folder AS image_folder,
                mm.name AS image_name,
                mm.path AS image_path,
                mm.module_name AS image_module_name,
                mm.reference_id AS image_reference_id,
                mm.file_type AS image_file_type,
                mm.file_size AS image_file_size,
                mm.created_at AS image_created_at,
                mm.updated_at AS image_updated_at
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
            LEFT JOIN (
                SELECT m1.*
                FROM media_manager m1
                INNER JOIN (
                    SELECT reference_id, MAX(id) AS max_id
                    FROM media_manager
                    WHERE folder IN ('marketing', 'announcement')
                    GROUP BY reference_id
                ) m2 ON m1.id = m2.max_id
            ) mm ON mm.reference_id = m.id
        """

        count_query = """
            SELECT COUNT(DISTINCT m.id) as total_count
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
            LEFT JOIN (
                SELECT reference_id, MAX(id) AS max_id
                FROM media_manager
                WHERE folder IN ('marketing', 'announcement')
                GROUP BY reference_id
            ) mm ON mm.reference_id = m.id
        """

        filters = []
        params = {}

        if id is not None:
            filters.append("m.id = :id")
            params["id"] = id

        if status_id is not None:
            filters.append("m.status_id = :status_id")
            params["status_id"] = status_id

        if marketing_type_id is not None:
            filters.append("m.marketing_type_id = :marketing_type_id")
            params["marketing_type_id"] = marketing_type_id

        if start_date:
            filters.append("DATE(m.created_at) >= :start_date")
            params["start_date"] = start_date

        if end_date:
            filters.append("DATE(m.created_at) <= :end_date")
            params["end_date"] = end_date

        if filters:
            where_clause = " WHERE " + " AND ".join(filters)
            base_query += where_clause
            count_query += where_clause

        total_count = db.execute(text(count_query), params).scalar()

        if id:  
            query = db.execute(text(base_query), params)
            raw_results = query.mappings().all()
            results = [format_marketing_record(dict(r)) for r in raw_results]
            return results, total_count

        offset = (page - 1) * record_count
        base_query += " ORDER BY m.id DESC LIMIT :limit OFFSET :offset"
        params["limit"] = record_count
        params["offset"] = offset

        query = db.execute(text(base_query), params)
        raw_results = query.mappings().all()
        results = [format_marketing_record(dict(r)) for r in raw_results]

        return results, total_count

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def list_announcements_service(
    db: Session,
    id: int = None,
    status_id: int = None,
    marketing_type_id: int = None,
    announcement_type: int = None,
    announcement_type_all: bool = False,
    start_date: datetime = None,
    end_date: datetime = None,
    page: int = 1,
    record_count: int = 10
):
    try:
        base_query = """
            SELECT 
                m.*,
                ms.name AS status_name,
                mt.key AS marketing_type_name,
                mm.id AS image_id,
                mm.folder AS image_folder,
                mm.name AS image_name,
                mm.path AS image_path,
                mm.module_name AS image_module_name,
                mm.reference_id AS image_reference_id,
                mm.file_type AS image_file_type,
                mm.file_size AS image_file_size,
                mm.created_at AS image_created_at,
                mm.updated_at AS image_updated_at
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
            LEFT JOIN (
                SELECT m1.*
                FROM media_manager m1
                INNER JOIN (
                    SELECT reference_id, MAX(id) AS max_id
                    FROM media_manager
                    WHERE folder = 'announcement'
                    GROUP BY reference_id
                ) m2 ON m1.id = m2.max_id
            ) mm ON mm.reference_id = m.id
        """

        count_query = """
            SELECT COUNT(DISTINCT m.id) as total_count
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
            LEFT JOIN (
                SELECT reference_id, MAX(id) AS max_id
                FROM media_manager
                WHERE folder = 'announcement'
                GROUP BY reference_id
            ) mm ON mm.reference_id = m.id
        """

        filters = []
        params = {}

        filters.append("m.announcement_type IS NOT NULL")

        if id is not None:
            filters.append("m.id = :id")
            params["id"] = id

        if status_id is not None:
            filters.append("m.status_id = :status_id")
            params["status_id"] = status_id

        if marketing_type_id is not None:
            filters.append("m.marketing_type_id = :marketing_type_id")
            params["marketing_type_id"] = marketing_type_id

        if not announcement_type_all and announcement_type is not None:
            filters.append("m.announcement_type = :announcement_type")
            params["announcement_type"] = announcement_type

        if start_date:
            filters.append("DATE(m.created_at) >= :start_date")
            params["start_date"] = start_date

        if end_date:
            filters.append("DATE(m.created_at) <= :end_date")
            params["end_date"] = end_date

        if filters:
            where_clause = " WHERE " + " AND ".join(filters)
            base_query += where_clause
            count_query += where_clause

        total_count = db.execute(text(count_query), params).scalar()

        if id:  
            query = db.execute(text(base_query), params)
            raw_results = query.mappings().all()
            results = [format_announcement_record(dict(r)) for r in raw_results]
            return results, total_count

        offset = (page - 1) * record_count
        base_query += " ORDER BY m.id DESC LIMIT :limit OFFSET :offset"
        params["limit"] = record_count
        params["offset"] = offset

        query = db.execute(text(base_query), params)
        raw_results = query.mappings().all()
        results = [format_announcement_record(dict(r)) for r in raw_results]

        return results, total_count

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def generate_excel_response(data, filename="announcements.xlsx"):
    """Generate Excel file response using OpenPyXL"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Announcements"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "S.No", "Title", "Description", "Status", "Start Date", "End Date"
        ]
        
        field_mapping = {
            "Title": "title",
            "Description": "description",
            "Status": "status_name",
            "Start Date": "start_date",
            "End Date": "end_date",
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for row_index, record in enumerate(data):
            row_num = row_index + 2  
            for col_num, header in enumerate(headers, 1):
                if header == "S.No":
                    value = row_index + 1
                else:
                    field_name = field_mapping.get(header)
                    value = record.get(field_name, "") if field_name else ""
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    if value is None or value == "":
                        value = "-"
                
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = cell_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(data) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50) 
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = "A2"

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

def generate_pdf_response(data, filename="announcements.pdf"):
    """Generate PDF file response using ReportLab"""
    try:
        pdf_buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        styles = getSampleStyleSheet()

        logo_url = get_logo_url_prod()


        try:
            logo = Image(logo_url)
            logo.drawHeight = 1.0 * inch
            logo.drawWidth = 2.0 * inch
            logo.hAlign = 'CENTER' 
            elements.append(logo)
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

        
        title_style = styles['Heading1']
        title_style.alignment = 1  
        title_style.fontSize = 16
        title_style.spaceAfter = 0.2*inch
        
        title = Paragraph("Announcements Report", title_style)
        elements.append(title)
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            alignment=1,  
            fontSize=10,
            textColor=colors.gray,
            spaceAfter=0.2*inch
        )
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
        elements.append(date_text)
        elements.append(Spacer(1, 0.2*inch))
        
        headers = [
            "S.No", "Title", "Description", "Status", "Start Date", "End Date"
        ]
        
        page_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
        col_widths = [
            page_width * 0.07,  
            page_width * 0.20,  
            page_width * 0.33,  
            page_width * 0.15,  
            page_width * 0.125, 
            page_width * 0.125, 
        ]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Headers centered
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ])

        max_rows_per_page = 10
        total_records = len(data)

        for page_index, start in enumerate(range(0, total_records, max_rows_per_page)):
            end = start + max_rows_per_page
            chunk = data[start:end]

            table_data = [headers]

            for idx, record in enumerate(chunk, start=start + 1):
                title = record.get('title', '') or "-"
                description = strip_html_tags(record.get('description', '')) or "-"
                status_name = record.get('status_name', '') or "-"
                start_d = format_date(record.get('start_date')) or "-"
                end_d = format_date(record.get('end_date')) or "-"

                row = [
                    str(idx),
                    title,
                    description,
                    status_name,
                    start_d,
                    end_d,
                ]
                table_data.append(row)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(table_style)
            elements.append(table)

            if end < total_records:
                elements.append(PageBreak())
        
        elements.append(Spacer(1, 0.2*inch))
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.darkblue
        )
        summary_text = Paragraph(f"Total Announcements: {len(data)}", summary_style)
        elements.append(summary_text)
        
        doc.build(elements)
        
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")

def to_iso_z(dt):
    if not dt:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return dt
    return dt.strftime("%Y-%m-%dT%H:%M:%S.%fZ")

def strip_html_tags(text: str) -> str:
    """Remove HTML tags from a string."""
    if not text:
        return None
    clean = re.sub(r'<.*?>', '', text)
    return clean.strip()
    
def new_list_marketing_service(
    db: Session,
    id: int = None,  
    status_id: int = None,
    marketing_type_id: int = None,
    page: int = 1,
    record_count: int = 10
):
    try:
        base_query = """
            SELECT 
                m.*,
                ms.id AS status_id,
                ms.key AS status_key,
                ms.name AS status_name,
                mt.id AS marketing_type_id,
                mt.name AS marketing_type_name
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
        """

        count_query = """
            SELECT COUNT(*) as total_count
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
        """

        filters = []
        params = {}

        if id is not None:
            filters.append("m.id = :id")
            params["id"] = id

        if status_id is not None:
            filters.append("m.status_id = :status_id")
            params["status_id"] = status_id

        if marketing_type_id is not None:
            filters.append("m.marketing_type_id = :marketing_type_id")
            params["marketing_type_id"] = marketing_type_id

        if filters:
            where_clause = " WHERE " + " AND ".join(filters)
            base_query += where_clause
            count_query += where_clause

        total_count = db.execute(text(count_query), params).scalar()

        offset = (page - 1) * record_count
        base_query += " ORDER BY m.id DESC LIMIT :limit OFFSET :offset"
        params["limit"] = record_count
        params["offset"] = offset

        query = db.execute(text(base_query), params).mappings().all()

        results = []
        for row in query:
            cover_image_path = row.get("cover_image")
            image_obj = None
            if cover_image_path:
                image_name = os.path.basename(cover_image_path)
                image_obj = {
                    "id": None,
                    "folder": "marketing",
                    "name": image_name,
                    "path": cover_image_path,
                    "module_name": image_name,
                    "reference_id": row["id"],
                    "file_type": os.path.splitext(image_name)[1][1:],
                    "created_at": str(row.get("created_at")),
                    "updated_at": str(row.get("updated_at")),
                    "deleted_at": None
                }

            results.append({
                "id": row["id"],
                "uuid": row.get("uuid"),
                "marketing_type_id": row.get("marketing_type_id"),
                "property_id": row.get("property_id"),
                "status_id": row.get("status_id"),
                "common_area_id": row.get("common_area_id"),
                "address": row.get("address"),
                "phone": row.get("phone"),
                "country_code": row.get("country_code"),
                "email": row.get("email"),
                "title": row.get("title"),
                "subtext": row.get("subtext"),
                "description": row.get("description"),
                "tags_removed_terms_condition": strip_html_tags(row.get("description")),
                "duration_start_date": str(row.get("duration_start_date")),
                "duration_end_date": str(row.get("duration_end_date")),
                "duration_from_time": str(row.get("duration_from_time")),
                "duration_end_time": str(row.get("duration_end_time")),
                "location_name": row.get("location_name"),
                "map_link": row.get("map_link"),
                "website": row.get("website"),
                # "terms_condition": row.get("terms_condition"),
                # "tags_removed_terms_condition": strip_html_tags(row.get("terms_condition")),
                "cover_image": cover_image_path,
                "start_date": to_iso_z(row.get("start_date")),
                "end_date": to_iso_z(row.get("end_date")),
                "created_by": row.get("created_by"),
                "images": image_obj,
                "status": {
                    "id": row.get("status_id"),
                    "key": row.get("status_key"),
                    "name": row.get("status_name")
                }
            })

        return results, total_count

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
def to_iso_z(value):
    if value:
        if isinstance(value, str):
            value = datetime.fromisoformat(value)
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    return None

def new_listing_marketing_service(
    db: Session,
    id: int = None,
    status_id: int = None,
    marketing_type_id: int = None,
    page: int = 1,
    record_count: int = 10
):
    try:
        base_query = """
            SELECT 
                m.*,
                ms.id AS status_id,
                ms.key AS status_key,
                ms.name AS status_name,
                ms.description AS status_description,
                ms.created_at AS status_created_at,
                ms.updated_at AS status_updated_at,
                mt.id AS marketing_type_id,
                mt.key AS marketing_type_key,
                mt.name AS marketing_type_name,
                mt.description AS marketing_type_description,
                mt.created_at AS marketing_type_created_at,
                mt.updated_at AS marketing_type_updated_at
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
        """

        count_query = """
            SELECT COUNT(*) as total_count
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
            LEFT JOIN marketing_type mt ON m.marketing_type_id = mt.id
        """

        filters = []
        params = {}

        if id is not None:
            filters.append("m.id = :id")
            params["id"] = id
        if status_id is not None:
            filters.append("m.status_id = :status_id")
            params["status_id"] = status_id
        if marketing_type_id is not None:
            filters.append("m.marketing_type_id = :marketing_type_id")
            params["marketing_type_id"] = marketing_type_id

        if filters:
            where_clause = " WHERE " + " AND ".join(filters)
            base_query += where_clause
            count_query += where_clause

        total_count = db.execute(text(count_query), params).scalar()

        offset = (page - 1) * record_count
        base_query += " ORDER BY m.id DESC LIMIT :limit OFFSET :offset"
        params["limit"] = record_count
        params["offset"] = offset

        row = db.execute(text(base_query), params).mappings().first()
        if not row:
            return None, total_count

        cover_image_path = row.get("cover_image")
        image_obj = None
        if cover_image_path:
            image_name = os.path.basename(cover_image_path)
            image_obj = {
                "id": None,
                "folder": "marketing",
                "name": image_name,
                "path": cover_image_path,
                "module_name": image_name,
                "reference_id": row["id"],
                "file_type": os.path.splitext(image_name)[1][1:],
                "created_at": to_iso_z(row.get("created_at")),
                "updated_at": to_iso_z(row.get("updated_at")),
                "deleted_at": None
            }

        result = {
            "id": row["id"],
            "uuid": row.get("uuid"),
            "marketing_type_id": row.get("marketing_type_id"),
            "property_id": row.get("property_id"),
            "status_id": row.get("status_id"),
            "common_area_id": row.get("common_area_id"),
            "address": row.get("address"),
            "phone": row.get("phone"),
            "country_code": row.get("country_code"),
            "email": row.get("email"),
            "title": row.get("title"),
            "subtext": row.get("subtext"),
            "description": row.get("description"),
            "duration_start_date": str(row.get("duration_start_date")),
            "duration_end_date": str(row.get("duration_end_date")),
            "duration_from_time": str(row.get("duration_from_time")),
            "duration_end_time": str(row.get("duration_end_time")),
            "location_name": row.get("location_name"),
            "map_link": row.get("map_link"),
            "website": row.get("website"),
            "terms_condition": row.get("terms_condition"),
            "cover_image": cover_image_path,
            "start_date": to_iso_z(row.get("start_date")),
            "end_date": to_iso_z(row.get("end_date")),
            "created_by": row.get("created_by"),
            "created_at": to_iso_z(row.get("created_at")),
            "updated_at": to_iso_z(row.get("updated_at")),
            "images": image_obj,
            "status": {
                "id": row.get("status_id"),
                "key": row.get("status_key"),
                "name": row.get("status_name"),
                "description": row.get("status_description"),
                "created_at": to_iso_z(row.get("status_created_at")),
                "updated_at": to_iso_z(row.get("status_updated_at"))
            },
            "marketing_type": {
                "id": row.get("marketing_type_id"),
                "key": row.get("marketing_type_key"),
                "name": row.get("marketing_type_name"),
                "description": row.get("marketing_type_description"),
                "created_at": to_iso_z(row.get("marketing_type_created_at")),
                "updated_at": to_iso_z(row.get("marketing_type_updated_at"))
            }
        }

        return result, total_count

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def list_marketing_search(
    db: Session,
    id: int = None,
    status_id: int = None,
    marketing_type_id: int = None,
    search: str = None,
    page: int = 1,
    record_count: int = 10
):
    try:
        base_query = """
            SELECT 
                m.*,
                ms.id AS status_id,
                ms.key AS status_key,
                ms.name AS status_name,
                ms.description AS status_description,
                ms.created_at AS status_created_at,
                ms.updated_at AS status_updated_at
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
        """

        count_query = """
            SELECT COUNT(*) as total_count
            FROM marketing m
            LEFT JOIN marketing_status ms ON m.status_id = ms.id
        """

        filters = []
        params = {}

        if id is not None:
            filters.append("m.id = :id")
            params["id"] = id
        if status_id is not None:
            filters.append("m.status_id = :status_id")
            params["status_id"] = status_id
        if marketing_type_id is not None:
            filters.append("m.marketing_type_id = :marketing_type_id")
            params["marketing_type_id"] = marketing_type_id
        if search:
            filters.append("""(
                LOWER(m.title) LIKE :search OR
                LOWER(m.address) LIKE :search OR
                LOWER(m.phone) LIKE :search OR
                LOWER(m.description) LIKE :search OR
                LOWER(m.location_name) LIKE :search
            )""")
            params["search"] = f"%{search.lower()}%"

        if filters:
            where_clause = " WHERE " + " AND ".join(filters)
            base_query += where_clause
            count_query += where_clause

        total_count = db.execute(text(count_query), params).scalar()

        offset = (page - 1) * record_count
        base_query += " ORDER BY m.id DESC LIMIT :limit OFFSET :offset"
        params["limit"] = record_count
        params["offset"] = offset

        rows = db.execute(text(base_query), params).mappings().all()
        results = []

        for row in rows:
            cover_image_path = row.get("cover_image")
            image_obj = None
            if cover_image_path:
                image_name = os.path.basename(cover_image_path)
                image_obj = {
                    "id": row.get("image_id", None),
                    "folder": "marketing",
                    "name": image_name,
                    "path": cover_image_path,
                    "module_name": image_name,
                    "reference_id": row["id"],
                    "file_type": os.path.splitext(image_name)[1][1:],
                    "created_at": to_iso_z(row.get("created_at")),
                    "updated_at": to_iso_z(row.get("updated_at")),
                    "deleted_at": None
                }

            results.append({
                "id": row["id"],
                "uuid": row.get("uuid"),
                "marketing_type_id": row.get("marketing_type_id"),
                "property_id": row.get("property_id"),
                "status_id": row.get("status_id"),
                "common_area_id": row.get("common_area_id"),
                "address": row.get("address"),
                "phone": row.get("phone"),
                "country_code": row.get("country_code"),
                "email": row.get("email"),
                "title": row.get("title"),
                "subtext": row.get("subtext"),
                "description": row.get("description"),
                "tags_removed_description": None,
                "duration_start_date": str(row.get("duration_start_date")),
                "duration_end_date": str(row.get("duration_end_date")),
                "duration_from_time": str(row.get("duration_from_time")),
                "duration_end_time": str(row.get("duration_end_time")),
                "location_name": row.get("location_name"),
                "map_link": row.get("map_link"),
                "website": row.get("website"),
                "terms_condition": row.get("terms_condition"),
                "tags_removed_terms_condition": None,
                "cover_image": cover_image_path,
                "start_date": to_iso_z(row.get("start_date")),
                "end_date": to_iso_z(row.get("end_date")),
                "created_by": row.get("created_by"),
                "images": image_obj,
                "status": {
                    "id": row.get("status_id"),
                    "key": row.get("status_key"),
                    "name": row.get("status_name"),
                    "description": row.get("status_description"),
                    "created_at": to_iso_z(row.get("status_created_at")),
                    "updated_at": to_iso_z(row.get("status_updated_at"))
                }
            })

        return results, total_count

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





def delete_marketing_service(db: Session, marketing_id: int):
    try:
        marketing = db.execute(
            text("SELECT * FROM marketing WHERE id = :id"),
            {"id": marketing_id}
        ).mappings().first()

        if not marketing:
            raise HTTPException(status_code=404, detail="Marketing record not found")

        db.execute(text("DELETE FROM marketing WHERE id = :id"), {"id": marketing_id})
        db.commit()

        return {
            "status": 200,
            "message": "Marketing deleted successfully."
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def update_announcement_active_status_service(
    db: Session,
    marketing_id: int,
    is_announcement_active: bool,
    status_id: int = None  
):
    """Update the active status and/or status_id of an announcement"""
    try:
        marketing = db.execute(
            text("SELECT * FROM marketing WHERE id = :id"),
            {"id": marketing_id}
        ).mappings().first()

        if not marketing:
            raise HTTPException(status_code=404, detail="Marketing record not found")

        if marketing.get("announcement_type") is None:
            raise HTTPException(
                status_code=400, 
                detail="This marketing record is not an announcement. Cannot update announcement status."
            )

        update_fields = ["is_announcement_active = :is_active", "updated_at = NOW()"]
        params = {
            "id": marketing_id,
            "is_active": 1 if is_announcement_active else 0
        }
        
        if status_id is not None:
            update_fields.append("status_id = :status_id")
            params["status_id"] = status_id
        
        update_sql = text(f"""
            UPDATE marketing 
            SET {', '.join(update_fields)}
            WHERE id = :id
        """)

        db.execute(update_sql, params)
        db.commit()

        updated_marketing = db.execute(
            text("SELECT * FROM marketing WHERE id = :id"),
            {"id": marketing_id}
        ).mappings().first()

        response_data = {
            "id": updated_marketing["id"],
            "title": updated_marketing["title"],
            "announcement_type": updated_marketing.get("announcement_type"),
            "is_announcement_active": bool(updated_marketing.get("is_announcement_active", False)),
            "status": "active" if is_announcement_active else "inactive"
        }
        
        if status_id is not None:
            response_data["status_id"] = updated_marketing.get("status_id")
        
        message = f"Announcement {'activated' if is_announcement_active else 'deactivated'}"
        if status_id is not None:
            message += f" and status updated to {status_id}"
        message += " successfully."

        return {
            "status": 200,
            "message": message,
            "data": response_data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))