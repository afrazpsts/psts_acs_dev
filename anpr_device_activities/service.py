from fastapi import APIRouter, Depends, Query, Request, UploadFile, File, HTTPException
import os
import json
import xmltodict
from sqlalchemy import text
from sqlalchemy.orm import Session  
from DB.db import get_db
from typing import List
import math
from typing import Optional
from common.logger import log as write_to_server_log
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from fastapi.responses import StreamingResponse
from utils.common_function import get_logo_url_prod, format_date
from datetime import datetime, timezone


router = APIRouter()

@router.post("/anpr_device_events")
async def hikvision_events(request: Request, db: Session = Depends(get_db)):
    headers_dict = dict(request.headers)
    write_to_server_log(f"ANPR Device Event - Headers: {json.dumps(headers_dict, indent=2, default=str)}")
    
    event_data = {
        "plate_number": None,
        "list_type": None,
        "gate_status": None,
        "xml_json": None,
        "dateTime": None,
        "country": None,
        "db_action": None,
        "channel_name": None,
        "barrier_gate_ctrl_type": None,
        "resident_id": None,
        "record_id": None
    }

    content_type = request.headers.get("content-type", "").lower()
    write_to_server_log(f"ANPR Device Event - Content-Type: {content_type}")

    if "application/json" in content_type:
        try:
            json_body = await request.json()
            write_to_server_log(f"ANPR Device Event - Raw JSON payload: {json.dumps(json_body, indent=2, default=str)}")
            
            xml_dict = json_body  
            event_data["xml_json"] = xml_dict
            
            event_notification = xml_dict.get("EventNotificationAlert", {})
            anpr = event_notification.get("ANPR", {})
            
            event_data["plate_number"] = anpr.get("licensePlate")
            event_data["dateTime"] = event_notification.get("dateTime")
            event_data["country"] = anpr.get("country")
            event_data["channel_name"] = event_notification.get("channelName")
            event_data["barrier_gate_ctrl_type"] = anpr.get("barrierGateCtrlType")

            raw_list_type = (anpr.get("listType") or "").strip().lower()
            if raw_list_type in ["allowlist", "white"]:
                event_data["list_type"] = "allowList"
            elif raw_list_type in ["blocklist", "black"]:
                event_data["list_type"] = "blockList"
            else:
                event_data["list_type"] = "unknown"

            write_to_server_log(f"ANPR Device Event - Successfully parsed JSON for plate: {event_data['plate_number']}")
            
        except Exception as e:
            error_msg = f"JSON parsing error: {str(e)}"
            event_data["xml_json"] = {"error": error_msg}
            write_to_server_log(f"ANPR Device Event - ERROR: {error_msg}")
    
    else:  
        try:
            form = await request.form()
            write_to_server_log(f"ANPR Device Event - Form data received: {dict(form)}")
            
            xml_file: UploadFile = form.get("anpr.xml")
            
            if xml_file:
                write_to_server_log(f"ANPR Device Event - XML file received: {xml_file.filename}")
                xml_bytes = await xml_file.read()
                
                xml_sample = xml_bytes[:500].decode('utf-8', errors='ignore')
                write_to_server_log(f"ANPR Device Event - XML sample: {xml_sample}...")
                
                try:
                    xml_dict = xmltodict.parse(xml_bytes)
                    write_to_server_log(f"ANPR Device Event - Parsed XML: {json.dumps(xml_dict, indent=2, default=str)}")
                    
                    event_data["xml_json"] = xml_dict

                    event_notification = xml_dict.get("EventNotificationAlert", {})
                    anpr = event_notification.get("ANPR", {})

                    event_data["plate_number"] = anpr.get("licensePlate")
                    event_data["dateTime"] = event_notification.get("dateTime")
                    event_data["country"] = anpr.get("country")
                    event_data["channel_name"] = event_notification.get("channelName")
                    event_data["barrier_gate_ctrl_type"] = anpr.get("barrierGateCtrlType")

                    raw_list_type = (anpr.get("listType") or "").strip().lower()
                    if raw_list_type in ["allowlist", "white"]:
                        event_data["list_type"] = "allowList"
                    elif raw_list_type in ["blocklist", "black"]:
                        event_data["list_type"] = "blockList"
                    else:
                        event_data["list_type"] = "unknown"

                    write_to_server_log(f"ANPR Device Event - Successfully parsed XML for plate: {event_data['plate_number']}")

                except Exception as e:
                    error_msg = f"XML parsing error: {str(e)}"
                    event_data["xml_json"] = {"error": error_msg}
                    write_to_server_log(f"ANPR Device Event - ERROR: {error_msg}")
        except Exception as e:
            error_msg = f"Form data error: {str(e)}"
            event_data["xml_json"] = {"error": error_msg}
            write_to_server_log(f"ANPR Device Event - ERROR: {error_msg}")

    plate_number = event_data["plate_number"]
    if plate_number:
        write_to_server_log(f"ANPR Device Event - Processing plate: {plate_number}, Channel: {event_data['channel_name']}, CtrlType: {event_data['barrier_gate_ctrl_type']}")
        
        plate_info = db.execute(
            text("""
                SELECT id, resident_id, listType, createTime, effectiveTime 
                FROM license_plate_access 
                WHERE LicensePlate = :plate_number
            """),
            {"plate_number": plate_number}
        ).first()

        if plate_info:
            current_time = datetime.now()
            is_time_valid = True
            time_validation_message = ""
            
            # Get resident_id as integer
            resident_id_val = int(plate_info.resident_id) if plate_info.resident_id else None
            
            if plate_info.createTime and plate_info.effectiveTime:
                try:
                    create_time_str = plate_info.createTime.replace('Z', '+00:00')
                    effective_time_str = plate_info.effectiveTime.replace('Z', '+00:00')
                    
                    create_time = datetime.fromisoformat(create_time_str)
                    effective_time = datetime.fromisoformat(effective_time_str)
                    
                    if current_time < create_time:
                        is_time_valid = False
                        time_validation_message = f"before valid start time ({plate_info.createTime})"
                    elif current_time > effective_time:
                        is_time_valid = False
                        time_validation_message = f"after valid end time ({plate_info.effectiveTime})"
                except Exception as e:
                    write_to_server_log(f"ANPR Device Event - Error parsing dates: {str(e)}")
            
            list_type_match = (event_data["list_type"].lower() == plate_info.listType.lower())
            write_to_server_log(f"ANPR Device Event - List type match: {list_type_match} (Device: {event_data['list_type']}, DB: {plate_info.listType})")
            
            channel_name = event_data["channel_name"] or ""
            is_entry = "Entry" in channel_name
            is_exit = "Exit" in channel_name
            
            should_open_gate = event_data["barrier_gate_ctrl_type"] == "1"
            write_to_server_log(f"ANPR Device Event - Should open gate: {should_open_gate} (CtrlType: {event_data['barrier_gate_ctrl_type']})")
            
            # FIRST: Check for open session (entry without exit) for this resident
            # Include resident_id in the SELECT
            open_session = db.execute(
                text("""
                    SELECT id, resident_id, entry_time 
                    FROM anpr_device_activities 
                    WHERE resident_id = :resident_id 
                    AND entry_time IS NOT NULL 
                    AND entry_time != ''
                    AND (exit_time IS NULL OR exit_time = '')
                    ORDER BY id DESC 
                    LIMIT 1
                """),
                {"resident_id": str(resident_id_val)}
            ).first()
            
            if open_session:
                write_to_server_log(f"ANPR Device Event - Found open session ID {open_session.id} for resident {open_session.resident_id} with entry at {open_session.entry_time}")
            
            # CRITICAL FIX: If it's an entry and there's an open session, DENY entry immediately
            if is_entry and open_session:
                write_to_server_log(f"ANPR Device Event - ENTRY BLOCKED: Resident {resident_id_val} already has open session (ID: {open_session.id})")
                
                payload = {
                    "plate_number": event_data["plate_number"],
                    "list_type": event_data["list_type"],
                    "gate_status": "DENIED - ALREADY INSIDE",
                    "reason": f"Vehicle already inside since {open_session.entry_time}",
                    "dateTime": event_data["dateTime"],
                    "country": event_data["country"],
                    "channel_name": event_data["channel_name"],
                    "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                }
                
                # Log the denied attempt
                db.execute(
                    text("""
                        INSERT INTO anpr_device_activities 
                        (anpr_device_activity, resident_id, created_at, updated_at)
                        VALUES (:json_data, :resident_id, :created_at, :updated_at)
                    """),
                    {
                        "json_data": json.dumps(payload),
                        "resident_id": str(resident_id_val),
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    }
                )
                db.commit()
                
                event_data["resident_id"] = resident_id_val
                event_data["db_action"] = f"Entry denied - already inside"
                event_data["gate_status"] = "NOT OPENED - VEHICLE ALREADY INSIDE"
                
                final_response = {"status": "ok", "event": event_data, "message": "Entry denied - vehicle already inside"}
                write_to_server_log(f"ANPR Device Event - Response: {json.dumps(final_response, indent=2, default=str)}")
                return final_response
            
            # Check for duplicate entry based on entry_time
            if is_entry and event_data["dateTime"]:
                # Check if there's already an entry with the same entry_time for this resident
                duplicate_entry = db.execute(
                    text("""
                        SELECT id, entry_time 
                        FROM anpr_device_activities 
                        WHERE resident_id = :resident_id 
                        AND entry_time = :entry_time
                        LIMIT 1
                    """),
                    {
                        "resident_id": str(resident_id_val),
                        "entry_time": event_data["dateTime"]
                    }
                ).first()
                
                if duplicate_entry:
                    write_to_server_log(f"ANPR Device Event - DUPLICATE ENTRY DETECTED: Same entry_time {event_data['dateTime']} for resident {resident_id_val}")
                    
                    event_data["resident_id"] = resident_id_val
                    event_data["db_action"] = f"Duplicate entry ignored (same timestamp)"
                    event_data["gate_status"] = "DUPLICATE ENTRY - IGNORED"
                    
                    final_response = {"status": "ok", "event": event_data, "message": "Duplicate entry ignored"}
                    write_to_server_log(f"ANPR Device Event - Response: {json.dumps(final_response, indent=2, default=str)}")
                    return final_response
            
            # Continue with normal processing only if no open session exists
            if not plate_info:
                payload = {
                    "plate_number": event_data["plate_number"],
                    "list_type": event_data["list_type"],
                    "gate_status": "DENIED - PLATE NOT FOUND",
                    "reason": "Plate not registered in system",
                    "dateTime": event_data["dateTime"],
                    "country": event_data["country"],
                    "channel_name": event_data["channel_name"],
                    "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                }
                
                db.execute(
                    text("""
                        INSERT INTO anpr_device_activities 
                        (anpr_device_activity, created_at, updated_at)
                        VALUES (:json_data, :created_at, :updated_at)
                    """),
                    {
                        "json_data": json.dumps(payload),
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    }
                )
                db.commit()
                result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                new_id = result.id if result else None
                
                event_data["db_action"] = f"Unknown plate attempt recorded"
                event_data["record_id"] = new_id
                event_data["gate_status"] = "NOT OPENED - PLATE NOT FOUND"
                write_to_server_log(f"ANPR Device Event - WARNING: Plate {plate_number} not found in database (ID: {new_id})")
                
            elif not is_time_valid:
                payload = {
                    "plate_number": event_data["plate_number"],
                    "list_type": event_data["list_type"],
                    "gate_status": "DENIED - TIME INVALID",
                    "reason": f"Access attempt {time_validation_message}",
                    "dateTime": event_data["dateTime"],
                    "country": event_data["country"],
                    "channel_name": event_data["channel_name"],
                    "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                }
                
                db.execute(
                    text("""
                        INSERT INTO anpr_device_activities 
                        (anpr_device_activity, resident_id, created_at, updated_at)
                        VALUES (:json_data, :resident_id, :created_at, :updated_at)
                    """),
                    {
                        "json_data": json.dumps(payload),
                        "resident_id": str(resident_id_val),
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    }
                )
                db.commit()
                result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                new_id = result.id if result else None
                
                event_data["resident_id"] = resident_id_val
                event_data["db_action"] = f"Time invalid attempt recorded"
                event_data["record_id"] = new_id
                event_data["gate_status"] = "NOT OPENED - TIME INVALID"
                
            elif not list_type_match:
                payload = {
                    "plate_number": event_data["plate_number"],
                    "list_type": event_data["list_type"],
                    "gate_status": "DENIED - LIST TYPE MISMATCH",
                    "reason": f"Device reports {event_data['list_type']}, DB has {plate_info.listType}",
                    "dateTime": event_data["dateTime"],
                    "country": event_data["country"],
                    "channel_name": event_data["channel_name"],
                    "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                }
                
                db.execute(
                    text("""
                        INSERT INTO anpr_device_activities 
                        (anpr_device_activity, resident_id, created_at, updated_at)
                        VALUES (:json_data, :resident_id, :created_at, :updated_at)
                    """),
                    {
                        "json_data": json.dumps(payload),
                        "resident_id": str(resident_id_val),
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    }
                )
                db.commit()
                result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                new_id = result.id if result else None
                
                event_data["resident_id"] = resident_id_val
                event_data["db_action"] = f"List type mismatch attempt recorded"
                event_data["record_id"] = new_id
                event_data["gate_status"] = "NOT OPENED - LIST TYPE MISMATCH"
                
            elif not should_open_gate:
                if is_entry:
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "GATE CLOSED - ENTRY ATTEMPT RECORDED",
                        "reason": "Gate close command received, entry not processed",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                elif is_exit:
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "GATE CLOSED - EXIT ATTEMPT RECORDED",
                        "reason": "Gate close command received, exit not processed",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                else:
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "GATE CLOSED",
                        "reason": "Gate close command received",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                
                db.execute(
                    text("""
                        INSERT INTO anpr_device_activities 
                        (anpr_device_activity, resident_id, created_at, updated_at)
                        VALUES (:json_data, :resident_id, :created_at, :updated_at)
                    """),
                    {
                        "json_data": json.dumps(payload),
                        "resident_id": str(resident_id_val),
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    }
                )
                db.commit()
                result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                new_id = result.id if result else None
                
                event_data["resident_id"] = resident_id_val
                event_data["db_action"] = f"Gate close command - no session update"
                event_data["record_id"] = new_id
                event_data["gate_status"] = "GATE CLOSED - NO ACCESS"
                write_to_server_log(f"ANPR Device Event - Gate closed for {plate_number}, no session update")
                
            elif list_type_match and is_time_valid and should_open_gate:
                if is_entry:
                    # ALLOW ENTRY - NO OPEN SESSION (already checked above)
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "OPENED - ENTRY",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                    
                    db.execute(
                        text("""
                            INSERT INTO anpr_device_activities 
                            (anpr_device_activity, resident_id, entry_time, created_at, updated_at)
                            VALUES (:json_data, :resident_id, :entry_time, :created_at, :updated_at)
                        """),
                        {
                            "json_data": json.dumps(payload),
                            "resident_id": str(resident_id_val),
                            "entry_time": event_data["dateTime"],
                            "created_at": datetime.now(),
                            "updated_at": datetime.now()
                        }
                    )
                    db.commit()
                    result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                    new_id = result.id if result else None
                    
                    event_data["resident_id"] = resident_id_val
                    event_data["db_action"] = f"Entry recorded"
                    event_data["record_id"] = new_id
                    event_data["gate_status"] = "OPENED - ENTRY"
                    write_to_server_log(f"ANPR Device Event - ENTRY ALLOWED for {plate_number} (ID: {new_id})")
                    
                elif is_exit and open_session:
                    # ALLOW EXIT
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "OPENED - EXIT",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                    
                    db.execute(
                        text("""
                            UPDATE anpr_device_activities 
                            SET exit_time = :exit_time, 
                                updated_at = :updated_at,
                                anpr_device_activity = :json_data
                            WHERE id = :session_id
                        """),
                        {
                            "session_id": open_session.id,
                            "exit_time": event_data["dateTime"],
                            "updated_at": datetime.now(),
                            "json_data": json.dumps(payload)
                        }
                    )
                    db.commit()
                    event_data["resident_id"] = open_session.resident_id
                    event_data["db_action"] = f"Exit recorded"
                    event_data["record_id"] = open_session.id
                    event_data["gate_status"] = "OPENED - EXIT"
                    write_to_server_log(f"ANPR Device Event - EXIT ALLOWED for {plate_number} (Session ID: {open_session.id})")
                    
                elif is_exit and not open_session:
                    # Exit with no entry - deny
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "DENIED - NOT INSIDE",
                        "reason": "No open entry session found",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                    
                    db.execute(
                        text("""
                            INSERT INTO anpr_device_activities 
                            (anpr_device_activity, resident_id, created_at, updated_at)
                            VALUES (:json_data, :resident_id, :created_at, :updated_at)
                        """),
                        {
                            "json_data": json.dumps(payload),
                            "resident_id": str(resident_id_val),
                            "created_at": datetime.now(),
                            "updated_at": datetime.now()
                        }
                    )
                    db.commit()
                    result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                    new_id = result.id if result else None
                    
                    event_data["resident_id"] = resident_id_val
                    event_data["db_action"] = f"Exit denied - not inside"
                    event_data["record_id"] = new_id
                    event_data["gate_status"] = "NOT OPENED - VEHICLE NOT INSIDE"
                else:
                    # Unknown channel
                    payload = {
                        "plate_number": event_data["plate_number"],
                        "list_type": event_data["list_type"],
                        "gate_status": "UNKNOWN CHANNEL",
                        "reason": f"Unrecognized channel: {channel_name}",
                        "dateTime": event_data["dateTime"],
                        "country": event_data["country"],
                        "channel_name": event_data["channel_name"],
                        "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"]
                    }
                    
                    db.execute(
                        text("""
                            INSERT INTO anpr_device_activities 
                            (anpr_device_activity, resident_id, created_at, updated_at)
                            VALUES (:json_data, :resident_id, :created_at, :updated_at)
                        """),
                        {
                            "json_data": json.dumps(payload),
                            "resident_id": str(resident_id_val),
                            "created_at": datetime.now(),
                            "updated_at": datetime.now()
                        }
                    )
                    db.commit()
                    result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
                    new_id = result.id if result else None
                    
                    event_data["resident_id"] = resident_id_val
                    event_data["db_action"] = f"Unknown channel recorded"
                    event_data["record_id"] = new_id
                    event_data["gate_status"] = "UNKNOWN"
        else:
            # Plate not found in database
            payload = {
                "plate_number": event_data["plate_number"],
                "list_type": event_data["list_type"],
                "gate_status": "DENIED - PLATE NOT FOUND",
                "dateTime": event_data["dateTime"],
                "country": event_data["country"],
                "channel_name": event_data["channel_name"],
                "barrier_gate_ctrl_type": event_data["barrier_gate_ctrl_type"],
                "reason": "Plate not registered in system"
            }
            
            db.execute(
                text("""
                    INSERT INTO anpr_device_activities 
                    (anpr_device_activity, created_at, updated_at)
                    VALUES (:json_data, :created_at, :updated_at)
                """),
                {
                    "json_data": json.dumps(payload),
                    "created_at": datetime.now(),
                    "updated_at": datetime.now()
                }
            )
            db.commit()
            result = db.execute(text("SELECT LAST_INSERT_ID() as id")).first()
            new_id = result.id if result else None
            
            event_data["db_action"] = f"Unknown plate attempt recorded"
            event_data["record_id"] = new_id
            event_data["gate_status"] = "NOT OPENED - PLATE NOT FOUND"
            write_to_server_log(f"ANPR Device Event - WARNING: Plate {plate_number} not found in database (ID: {new_id})")

    final_response = {"status": "ok", "event": event_data}
    
    write_to_server_log(f"ANPR Device Event - Final event data: {json.dumps(event_data, indent=2, default=str)}")
    write_to_server_log(f"ANPR Device Event - Response being sent: {json.dumps(final_response, indent=2, default=str)}")
    
    print(" Final Hikvision ANPR Event:", json.dumps(event_data, indent=2, default=str))

    return final_response

@router.get("/anpr_activities")
def list_anpr_activities(
    db: Session = Depends(get_db),
    from_date: Optional[str] = Query(None, description="Filter start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="Filter end date (YYYY-MM-DD)"),
    searchData: Optional[str] = Query(None, description="Search by plate number, resident name, card number, or resident type"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(10, ge=1, le=100, description="Records per page"),
    sort: str = Query("recent", regex="^(recent|old|null)$", description="Sort order: recent, old, or null"),
    show_all: bool = Query(False, description="Set to true to show all records including denied attempts"),
    download: bool = Query(False, description="Set to true to download as Excel or PDF"),
    download_type: Optional[str] = Query('excel', description="Download format: 'excel' or 'pdf'"),
):
    try:
        base_query = """
            SELECT 
                ada.id,
                ada.anpr_device_activity,
                ada.resident_id,
                ada.entry_time,
                ada.exit_time,
                ada.created_at as activity_created_at,
                ada.updated_at as activity_updated_at,
                
                -- User personal details
                up.id as user_id,
                up.first_name,
                up.last_name,
                up.email,
                up.phone,
                
                -- License plate access details
                lpa.id as license_plate_id,
                lpa.LicensePlate,
                lpa.iu_number,
                lpa.listType as db_list_type,
                lpa.vehicle_type,
                lpa.createTime as license_create_time,
                lpa.effectiveTime as license_effective_time,
                lpa.building_id,
                
                -- Level details (from building_level)
                bl.id as level_id,
                bl.level as level_name,
                
                -- Unit details (from building_units)
                bu.id as unit_id,
                bu.unit_no as unit_name,
                
                -- Residency type details
                rt.id as residency_type_id,
                rt.name as resident_type,
                
                -- User access details
                uad.join_date,
                uad.access_start,
                uad.access_end
                
            FROM anpr_device_activities ada
            LEFT JOIN license_plate_access lpa 
                ON ada.resident_id = lpa.resident_id
                AND JSON_UNQUOTE(JSON_EXTRACT(ada.anpr_device_activity, '$.plate_number')) = lpa.LicensePlate
            LEFT JOIN user_personal_details up 
                ON lpa.resident_id = up.id
            LEFT JOIN user_access_details uad
                ON lpa.resident_id = uad.user_id
            LEFT JOIN residency_type rt
                ON uad.residency_type_id = rt.id
            LEFT JOIN building_level bl
                ON uad.level_id = bl.id
            LEFT JOIN building_units bu
                ON uad.unit_id = bu.id
            WHERE 1=1
        """

        if not show_all:
            base_query += " AND ada.entry_time IS NOT NULL"

        filters = {}

        if from_date and from_date.lower() != "null":
            base_query += " AND DATE(ada.created_at) >= :from_date"
            filters["from_date"] = from_date

        if to_date and to_date.lower() != "null":
            base_query += " AND DATE(ada.created_at) <= :to_date"
            filters["to_date"] = to_date

        if searchData and searchData.lower() != "null":
            base_query += """
                AND (
                    LOWER(JSON_UNQUOTE(JSON_EXTRACT(ada.anpr_device_activity, '$.plate_number'))) LIKE :search
                    OR LOWER(JSON_UNQUOTE(JSON_EXTRACT(ada.anpr_device_activity, '$.list_type'))) LIKE :search
                    OR LOWER(JSON_UNQUOTE(JSON_EXTRACT(ada.anpr_device_activity, '$.gate_status'))) LIKE :search
                    OR LOWER(JSON_UNQUOTE(JSON_EXTRACT(ada.anpr_device_activity, '$.channel_name'))) LIKE :search
                    OR LOWER(up.first_name) LIKE :search
                    OR LOWER(up.last_name) LIKE :search
                    OR LOWER(lpa.iu_number ) LIKE :search
                    OR LOWER(lpa.LicensePlate) LIKE :search
                    OR LOWER(rt.name) LIKE :search
                    OR LOWER(bl.level) LIKE :search
                    OR LOWER(bu.unit_no) LIKE :search
                )
            """
            filters["search"] = f"%{searchData.lower()}%"

        if sort == "recent":
            order_by = "ada.created_at DESC"
        elif sort == "old":
            order_by = "ada.created_at ASC"
        else:
            order_by = "ada.created_at DESC"

        count_query = f"SELECT COUNT(*) as total FROM ({base_query}) as sub"
        total_records = db.execute(text(count_query), filters).scalar() or 0
        last_page = math.ceil(total_records / per_page) if total_records > 0 else 1

        if download:
            final_query = f"""
                {base_query}
                ORDER BY {order_by}
            """
        else:
            offset = (page - 1) * per_page
            final_query = f"""
                {base_query}
                ORDER BY {order_by}
                LIMIT :limit OFFSET :offset
            """
            filters.update({"limit": per_page, "offset": offset})

        rows = db.execute(text(final_query), filters).fetchall()

        results = []
        for row in rows:
            activity = json.loads(row.anpr_device_activity) if row.anpr_device_activity else {}
            
            building_name = None
            if row.building_id:
                building_name = f"Building {row.building_id}"
            
            parking_duration = None
            parking_duration_formatted = None
            parking_duration_days = None
            parking_duration_hours = None
            parking_status = None
            
            if row.entry_time:
                try:
                    entry_time_str = str(row.entry_time)
                    entry_time_str = entry_time_str.replace('Z', '+00:00')
                    
                    entry_time = datetime.fromisoformat(entry_time_str)
                    
                    if entry_time.tzinfo is not None:
                        entry_time = entry_time.astimezone(timezone.utc).replace(tzinfo=None)
                    
                    if row.exit_time and row.exit_time != "NULL" and row.exit_time:
                        exit_time_str = str(row.exit_time).replace('Z', '+00:00')
                        exit_time = datetime.fromisoformat(exit_time_str)
                        
                        if exit_time.tzinfo is not None:
                            exit_time = exit_time.astimezone(timezone.utc).replace(tzinfo=None)
                        
                        end_time = exit_time
                        status = "COMPLETED"
                    else:
                        end_time = datetime.now().replace(tzinfo=None)
                        status = "PARKED"
                    
                    duration = end_time - entry_time
                    
                    total_seconds = int(duration.total_seconds())
                    
                    if total_seconds < 0:
                        total_seconds = abs(total_seconds)
                    
                    total_hours = total_seconds / 3600
                    
                    days = int(total_hours // 24)
                    remaining_hours = int(total_hours % 24)
                    total_minutes = int(total_seconds / 60)
                    minutes = total_minutes % 60
                    
                    if days > 0:
                        if days == 1:
                            if remaining_hours > 0:
                                if remaining_hours == 1:
                                    parking_duration_formatted = f"1 day 1 hour"
                                else:
                                    parking_duration_formatted = f"1 day {remaining_hours} hours"
                            else:
                                parking_duration_formatted = f"1 day"
                        else:
                            if remaining_hours > 0:
                                if remaining_hours == 1:
                                    parking_duration_formatted = f"{days} days 1 hour"
                                else:
                                    parking_duration_formatted = f"{days} days {remaining_hours} hours"
                            else:
                                parking_duration_formatted = f"{days} days"
                    elif remaining_hours > 0:
                        if remaining_hours == 1:
                            if minutes > 0:
                                if minutes == 1:
                                    parking_duration_formatted = f"1 hour 1 minute"
                                else:
                                    parking_duration_formatted = f"1 hour {minutes} minutes"
                            else:
                                parking_duration_formatted = f"1 hour"
                        else:
                            if minutes > 0:
                                if minutes == 1:
                                    parking_duration_formatted = f"{remaining_hours} hours 1 minute"
                                else:
                                    parking_duration_formatted = f"{remaining_hours} hours {minutes} minutes"
                            else:
                                parking_duration_formatted = f"{remaining_hours} hours"
                    else:
                        if minutes == 1:
                            parking_duration_formatted = f"1 minute"
                        elif minutes > 0:
                            parking_duration_formatted = f"{minutes} minutes"
                        else:
                            parking_duration_formatted = "Less than a minute"
                    
                    parking_duration = {
                        "total_seconds": total_seconds,
                        "total_minutes": round(total_seconds / 60, 2),
                        "total_hours": round(total_hours, 2),
                        "days": days,
                        "hours": remaining_hours,
                        "minutes": minutes,
                        "status": status
                    }
                    
                    parking_duration_days = days
                    parking_duration_hours = remaining_hours
                    parking_status = status
                    
                except Exception as e:
                    write_to_server_log(f"Error calculating parking duration for entry_time {row.entry_time}: {str(e)}")
                    parking_duration = {
                        "error": str(e),
                        "entry_time": row.entry_time,
                        "status": "ERROR"
                    }
                    parking_duration_formatted = "Calculation error"
                    parking_duration_days = 0
                    parking_duration_hours = 0
                    parking_status = "ERROR"
            
            result = {
                "id": row.id,
                "resident_id": row.resident_id,
                "entry_time": row.entry_time,
                "exit_time": row.exit_time,
                "activity_created_at": row.activity_created_at.strftime("%Y-%m-%d %H:%M:%S") if row.activity_created_at else None,
                "activity_updated_at": row.activity_updated_at.strftime("%Y-%m-%d %H:%M:%S") if row.activity_updated_at else None,
                
                "parking_duration": parking_duration,
                "parking_duration_formatted": parking_duration_formatted,
                "parking_duration_days": parking_duration_days,
                "parking_duration_hours": parking_duration_hours,
                "parking_status": parking_status,
                
                "plate_number": activity.get("plate_number"),
                "list_type": activity.get("list_type"),
                "gate_status": activity.get("gate_status"),
                "dateTime": activity.get("dateTime"),
                "country": activity.get("country"),
                "channel_name": activity.get("channel_name"),
                "barrier_gate_ctrl_type": activity.get("barrier_gate_ctrl_type"),
                "reason": activity.get("reason") if activity.get("reason") else None,
                
                "resident_name": f"{row.first_name or ''} {row.last_name or ''}".strip() or None,
                "first_name": row.first_name,
                "last_name": row.last_name,
                "email": row.email,
                "phone": row.phone,
                
                "license_plate_id": row.license_plate_id,
                "LicensePlate": row.LicensePlate,
                "card_no": row.iu_number,
                "db_list_type": row.db_list_type,
                "vehicle_type": row.vehicle_type,
                "license_create_time": row.license_create_time,
                "license_effective_time": row.license_effective_time,
                
                "building_id": row.building_id,
                "building_name": building_name,
                
                "level_id": row.level_id,
                "level_name": row.level_name,
                
                "unit_id": row.unit_id,
                "unit_name": row.unit_name,
                
                "residency_type_id": row.residency_type_id,
                "resident_type": row.resident_type,
                
                "join_date": row.join_date,
                "access_start": row.access_start,
                "access_end": row.access_end,
            }
            
            results.append(result)

        if download:
            filename_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            if (download_type or "").lower() == "pdf":
                return generate_anpr_activities_pdf(results, f"anpr_activities_{filename_ts}.pdf")
            return generate_anpr_activities_excel(results, f"anpr_activities_{filename_ts}.xlsx")

        return {
            "status": 200,
            "message": "ANPR activities retrieved successfully",
            "data": results,
            "pagination_details": {
                "page": page,
                "per_page": per_page,
                "total": total_records,
                "last_page": last_page
            },
            "filter_mode": "all_records" if show_all else "only_entries_exits"
        }
        
    except Exception as e:
        write_to_server_log(f"Error in list_anpr_activities: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve ANPR activities: {str(e)}")


def generate_anpr_activities_excel(data: List[dict], filename: str = "anpr_activities.xlsx"):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ANPR Activities"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "S.No",
            "Resident Name",
            "Email",
            "Block",
            "Level-unit",
            "Plate Number",
            "Entry Time",
            "Exit Time",
            "Parking Duration",
            "Parking Status",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for row_index, record in enumerate(data):
            row_num = row_index + 2
            resident_name = record.get("resident_name") or "-"
            email = record.get("email") or "-"
            building_name = record.get("building_name") or "-"
            level_unit = None
            level_name = record.get("level_name") or ""
            unit_name = record.get("unit_name") or ""
            if level_name and unit_name:
                level_unit = f"{level_name}-{unit_name}"
            elif level_name:
                level_unit = level_name
            elif unit_name:
                level_unit = unit_name
            else:
                level_unit = "-"

            plate_number = record.get("plate_number") or record.get("LicensePlate") or "-"

            entry_time = record.get("entry_time")
            exit_time = record.get("exit_time")
            entry_time_str = str(entry_time) if entry_time else "-"
            exit_time_str = str(exit_time) if exit_time else "-"

            parking_duration_formatted = record.get("parking_duration_formatted") or "-"
            parking_status = record.get("parking_status") or "-"

            row_values = [
                row_index + 1,
                resident_name,
                email,
                building_name,
                level_unit,
                plate_number,
                entry_time_str,
                exit_time_str,
                parking_duration_formatted,
                parking_status,
            ]

            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = cell_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(data) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")


def generate_anpr_activities_pdf(data: List[dict], filename: str = "anpr_activities.pdf"):
    try:
        pdf_buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.3 * inch,
            leftMargin=0.3 * inch,
            topMargin=0.4 * inch,
            bottomMargin=0.4 * inch
        )

        elements = []
        styles = getSampleStyleSheet()

        logo_url = get_logo_url_prod()
        try:
            logo = Image(logo_url)
            logo.drawHeight = 0.8 * inch
            logo.drawWidth = 1.6 * inch
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

        title_style = styles['Heading1']
        title_style.alignment = 1
        title_style.fontSize = 14
        title_style.spaceAfter = 0.1 * inch

        elements.append(Paragraph("Parking Activities Report", title_style))

        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            alignment=1,
            fontSize=8,
            textColor=colors.gray,
            spaceAfter=0.1 * inch
        )

        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.1 * inch))

        headers = [
            "S.No",
            "Resident Name",
            "Email",
            "Block",
            "Level-Unit",
            "Plate Number",
            "Entry Time",
            "Exit Time",
            "Duration",
            "Status",
        ]

        page_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
        
        
        col_widths = [
            page_width * 0.05,  # S.No
            page_width * 0.12,  # Resident N
            page_width * 0.15,  
            page_width * 0.10, 
            page_width * 0.10,
            page_width * 0.10,  
            page_width * 0.12,  
            page_width * 0.12,  
            page_width * 0.08, 
            page_width * 0.06,  
        ]

        def format_datetime(dt_str):
            """Format datetime string to readable format"""
            if not dt_str or dt_str == "-":
                return "-"
            try:
            
                if "T" in str(dt_str):
                    parts = str(dt_str).split('T')
                    if len(parts) > 1:
                        time_part = parts[1].split('.')[0][:8]  # Get HH:MM:SS
                        return f"{parts[0]} {time_part}"
                
                dt_str_clean = str(dt_str)
                if 'T' in dt_str_clean:
                    date_part = dt_str_clean.split('T')[0]
                    time_part = dt_str_clean.split('T')[1].split('.')[0][:8] 
                    return f"{date_part} {time_part}"
                return dt_str_clean[:19]  
            except:
                return str(dt_str)[:19]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 6.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
        ])

        table_style.add('BACKGROUND', (9, 1), (9, 1), colors.HexColor('#E6F0FA'))  
        table_style.add('BACKGROUND', (9, 2), (9, 3), colors.HexColor('#E6FFE6')) 

        max_rows_per_page = 12  
        total_records = len(data)

        for start in range(0, total_records, max_rows_per_page):
            end = start + max_rows_per_page
            chunk = data[start:end]

            table_data = [headers]

            for idx, record in enumerate(chunk, start=start + 1):
                resident_name = record.get("resident_name") or "-"
                if len(resident_name) > 15:
                    resident_name = resident_name[:12] + "..."  # Truncate long names
                
                email = record.get("email") or "-"
                if len(email) > 20:
                    email = email[:18] + "..."  # Truncate long emails
                    
                building_name = record.get("building_name") or "-"
                
                level_name = record.get("level_name") or ""
                unit_name = record.get("unit_name") or ""
                if level_name and unit_name:
                    level_unit = f"{level_name}-{unit_name}"
                elif level_name:
                    level_unit = level_name
                elif unit_name:
                    level_unit = unit_name
                else:
                    level_unit = "-"

                plate_number = record.get("plate_number") or record.get("LicensePlate") or "-"

                entry_time = record.get("entry_time")
                exit_time = record.get("exit_time")
                entry_time_str = format_datetime(entry_time)
                exit_time_str = format_datetime(exit_time)

                parking_duration_formatted = record.get("parking_duration_formatted") or "-"
                parking_status = record.get("parking_status") or "-"

                row = [
                    str(idx),
                    resident_name,
                    email,
                    building_name,
                    level_unit,
                    plate_number,
                    entry_time_str,
                    exit_time_str,
                    parking_duration_formatted,
                    parking_status,
                ]

                table_data.append(row)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(table_style)
            elements.append(table)

            if end < total_records:
                elements.append(PageBreak())

        elements.append(Spacer(1, 0.1 * inch))
        
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#2F75B5'),
            alignment=2,  
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(f"Total ANPR Activities: {len(data)}", summary_style))

        doc.build(elements)
        pdf_buffer.seek(0)

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")