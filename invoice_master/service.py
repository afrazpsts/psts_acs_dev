from sqlalchemy import text, and_, or_
from sqlalchemy.exc import ProgrammingError
from sqlalchemy.orm import Session
from common.logger import log as write_to_server_log
from typing import Optional, List, Dict, Any
from datetime import datetime, date
import json
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from utils.common_function import get_logo_url_prod

from DB.db import SessionLocal


def generate_invoice_number(db: Session) -> str:
    """Generate unique invoice number like INV0034"""
    try:
        result = db.execute(
            text("SELECT invoice_number FROM invoice_master ORDER BY id DESC LIMIT 1")
        ).first()
        
        if result and result[0]:
            last_number = result[0]
            num_part = int(last_number.replace('INV', ''))
            new_num = num_part + 1
        else:
            new_num = 1
        
        return f"INV{str(new_num).zfill(4)}"
    except Exception as e:
        write_to_server_log(f"Error generating invoice number: {str(e)}")
        return f"INV{datetime.now().strftime('%Y%m%d%H%M%S')}"


def convert_period_to_uppercase(period: str) -> str:
    """Convert period to uppercase for enum compatibility"""
    period_map = {
        'day': 'DAY',
        'week': 'WEEK', 
        'month': 'MONTH',
        'year': 'YEAR'
    }
    return period_map.get(period.lower(), period.upper())


def convert_month_option_to_uppercase(option: str) -> str:
    """Convert month_option to uppercase for enum compatibility"""
    if not option:
        return None
    option_map = {
        'date': 'DATE',
        'day': 'DAY'
    }
    return option_map.get(option.lower(), option.upper())


def convert_end_type_to_uppercase(end_type: str) -> str:
    """Convert end_type to uppercase for enum compatibility"""
    if not end_type:
        return 'NEVER'
    end_type_map = {
        'never': 'NEVER',
        'on': 'ON',
        'after': 'AFTER'
    }
    return end_type_map.get(end_type.lower(), end_type.upper())


def create_invoice_service(invoice_data: Dict[str, Any], db: Session):
    """
    Service function to create a new invoice with vehicle items and recurring settings
    """
    try:
        # Filter vehicle items to only include checked=true items
        all_vehicle_items = invoice_data.get('vehicle_items', [])
        checked_vehicle_items = [item for item in all_vehicle_items if item.get('checked') == True]
        
        # Update invoice_data with filtered vehicle items
        invoice_data['vehicle_items'] = checked_vehicle_items
        
        # Validate that at least one vehicle item is checked
        if not checked_vehicle_items:
            raise Exception("At least one vehicle item must be selected (checked=true)")
        
        # Extract values from payload with proper defaults
        total_sub_total = float(invoice_data.get('sub_total', 0))
        total_gst = float(invoice_data.get('gst', 0))
        total_amount = float(invoice_data.get('total_amount', 0))
        discount = float(invoice_data.get('discount', 0))
        
        # Process extra_amount from invoice level (if any)
        invoice_level_extra_amount = []
        if 'extra_amount' in invoice_data and invoice_data.get('extra_amount'):
            invoice_extra = invoice_data.get('extra_amount')
            if isinstance(invoice_extra, list):
                invoice_level_extra_amount = invoice_extra
        
        # Generate invoice number
        invoice_number = generate_invoice_number(db)
        
        # Convert extra_amount to JSON string for storage
        extra_amount_json = json.dumps(invoice_level_extra_amount) if invoice_level_extra_amount else None
        
        # Insert invoice master
        insert_query = """
            INSERT INTO invoice_master (
                invoice_number, resident_id, building_id, level_id, unit_id,
                invoice_date, due_date, sub_total, extra_amount, gst, total_amount, discount,
                terms_and_conditions, status, payment_status, mark_as_recurring,
                created_by, created_at, updated_at
            ) VALUES (
                :invoice_number, :resident_id, :building_id, :level_id, :unit_id,
                :invoice_date, :due_date, :sub_total, :extra_amount, :gst, :total_amount, :discount,
                :terms_and_conditions, :status, :payment_status, :mark_as_recurring,
                :created_by, NOW(), NOW()
            )
        """
        
        db.execute(
            text(insert_query),
            {
                "invoice_number": invoice_number,
                "resident_id": int(invoice_data.get('resident_id')),
                "building_id": int(invoice_data.get('building_id')),
                "level_id": int(invoice_data.get('level_id')) if invoice_data.get('level_id') else None,
                "unit_id": int(invoice_data.get('unit_id')) if invoice_data.get('unit_id') else None,
                "invoice_date": invoice_data.get('invoice_date').split('T')[0] if invoice_data.get('invoice_date') else None,
                "due_date": invoice_data.get('due_date').split('T')[0] if invoice_data.get('due_date') else None,
                "sub_total": total_sub_total,
                "extra_amount": extra_amount_json,
                "gst": total_gst,
                "total_amount": total_amount,
                "discount": discount,
                "terms_and_conditions": invoice_data.get('terms_and_conditions', ''),
                "status": 'DRAFT',
                "payment_status": 'PENDING',
                "mark_as_recurring": 1 if invoice_data.get('mark_as_recurring') else 0,
                "created_by": invoice_data.get('created_by', 1)
            }
        )
        db.commit()
        
        # Get the inserted invoice ID
        result = db.execute(
            text("SELECT id FROM invoice_master WHERE invoice_number = :invoice_number"),
            {"invoice_number": invoice_number}
        ).first()
        
        invoice_id = result[0] if result else None
        
        if not invoice_id:
            raise Exception("Failed to get invoice ID after creation")
        
        # Insert only checked vehicle items
        for item in checked_vehicle_items:
            # Handle extraCharges from the payload
            extra_charges = item.get('extraCharges', [])
            extra_amount_json = json.dumps(extra_charges) if extra_charges else None
            
            # Calculate totals from item fields
            # Use calculated_amount if available, otherwise base_amount
            sub_total = float(item.get('calculated_amount', item.get('base_amount', 0)))
            
            # Get discount from item if available, otherwise 0
            item_discount = float(item.get('discount', 0)) if item.get('discount') else 0
            
            # For vehicle items, gst and total might be calculated at invoice level
            item_gst = 0  # Default, since GST is at invoice level in your payload
            item_total = sub_total
            
            # Prepare description with additional info
            description_parts = []
            if item.get('vehicle_owner_type'):
                description_parts.append(f"Owner type: {item.get('vehicle_owner_type')}")
            if item.get('billing_period'):
                description_parts.append(f"Billing period: {item.get('billing_period')}")
            if item.get('no_of_days') is not None:
                description_parts.append(f"Days: {item.get('no_of_days')}")
            if item.get('per_day_amount'):
                description_parts.append(f"Per day amount: {item.get('per_day_amount')}")
            
            description = ", ".join(description_parts) if description_parts else None
            
            vehicle_item_query = """
                INSERT INTO invoice_vehicle_items (
                    invoice_id, vehicle_id, iu_number, vehicle_number, vehicle_type_id,
                    sub_total, extra_amount, gst, total_amount, discount, description, created_by, created_at, updated_at
                ) VALUES (
                    :invoice_id, :vehicle_id, :iu_number, :vehicle_number, :vehicle_type_id,
                    :sub_total, :extra_amount, :gst, :total_amount, :discount, :description, :created_by, NOW(), NOW()
                )
            """
            
            db.execute(
                text(vehicle_item_query),
                {
                    "invoice_id": invoice_id,
                    "vehicle_id": item.get('vehicle_id'),
                    "iu_number": item.get('iu_number'),
                    "vehicle_number": item.get('vehicle_number'),
                    "vehicle_type_id": item.get('vehicle_type_id') or item.get('vehicle_type'),  # Try vehicle_type_id first, fallback to vehicle_type
                    "sub_total": sub_total,
                    "extra_amount": extra_amount_json,
                    "gst": item_gst,
                    "total_amount": item_total,
                    "discount": item_discount,
                    "description": description,
                    "created_by": invoice_data.get('created_by', 1)
                }
            )
        
        # Insert into invoice_recurring_residents if mark_as_recurring is True
        if invoice_data.get('mark_as_recurring'):
            recurring_resident_query = """
                INSERT INTO invoice_recurring_residents (
                    resident_id, invoice_id, building_id, level_id, unit_id, created_by, created_at, updated_at
                ) VALUES (
                    :resident_id, :invoice_id, :building_id, :level_id, :unit_id, :created_by, NOW(), NOW()
                )
            """
            
            db.execute(
                text(recurring_resident_query),
                {
                    "resident_id": int(invoice_data.get('resident_id')),
                    "invoice_id": invoice_id,
                    "building_id": int(invoice_data.get('building_id')),
                    "level_id": int(invoice_data.get('level_id')) if invoice_data.get('level_id') else None,
                    "unit_id": int(invoice_data.get('unit_id')) if invoice_data.get('unit_id') else None,
                    "created_by": invoice_data.get('created_by', 1)
                }
            )
        
        # Insert recurring settings if marked as recurring
        if invoice_data.get('mark_as_recurring') and 'period' in invoice_data:
            # Convert values to uppercase for enum compatibility
            period_upper = convert_period_to_uppercase(invoice_data.get('period', 'month'))
            month_option_upper = convert_month_option_to_uppercase(invoice_data.get('monthOption', 'date'))
            end_type_upper = convert_end_type_to_uppercase(invoice_data.get('endType', 'never'))
            
            # Use backticks for column names that might be reserved keywords
            recurring_query = """
                INSERT INTO recurring_invoice_settings (
                    invoice_id, repeat_every, period, month_option, day_of_month,
                    weekday, week_number, selected_days, `year_month`, `year_day`,
                    end_type, end_date, after_occurrences, next_invoice_date,
                    is_active, created_at, updated_at
                ) VALUES (
                    :invoice_id, :repeat_every, :period, :month_option, :day_of_month,
                    :weekday, :week_number, :selected_days, :year_month, :year_day,
                    :end_type, :end_date, :after_occurrences, :next_invoice_date,
                    :is_active, NOW(), NOW()
                )
            """
            
            next_date = calculate_next_recurring_date(invoice_data)
            
            # Handle empty string values
            end_date_value = invoice_data.get('endDate')
            if end_date_value == '':
                end_date_value = None
            
            after_occurrences_value = invoice_data.get('afterOccurrences')
            if after_occurrences_value == '':
                after_occurrences_value = None
            elif after_occurrences_value:
                after_occurrences_value = int(after_occurrences_value)
            
            db.execute(
                text(recurring_query),
                {
                    "invoice_id": invoice_id,
                    "repeat_every": invoice_data.get('repeatEvery', 1),
                    "period": period_upper,
                    "month_option": month_option_upper,
                    "day_of_month": invoice_data.get('dayOfMonth'),
                    "weekday": invoice_data.get('weekday'),
                    "week_number": invoice_data.get('weekNumber'),
                    "selected_days": json.dumps(invoice_data.get('selectedDays', [])) if invoice_data.get('selectedDays') else None,
                    "year_month": invoice_data.get('yearMonth'),
                    "year_day": invoice_data.get('yearDay'),
                    "end_type": end_type_upper,
                    "end_date": end_date_value,
                    "after_occurrences": after_occurrences_value,
                    "next_invoice_date": next_date,
                    "is_active": 1
                }
            )
        
        db.commit()
        
        # Return created invoice
        return get_invoice_by_id_service(invoice_id, db)
        
    except Exception as e:
        db.rollback()
        error_msg = str(e)
        write_to_server_log(f"Error in create_invoice_service: {error_msg}")
        import traceback
        write_to_server_log(f"Traceback: {traceback.format_exc()}")
        raise Exception(f"Failed to create invoice: {error_msg}")


def calculate_next_recurring_date(recurring_data: Dict[str, Any]) -> date:
    """Calculate next invoice date based on recurrence pattern"""
    from datetime import datetime, timedelta
    import calendar
    
    today = datetime.now().date()
    
    period = recurring_data.get('period', 'month')
    repeat_every = recurring_data.get('repeatEvery', 1)
    
    if period == 'day':
        return today + timedelta(days=repeat_every)
    elif period == 'week':
        return today + timedelta(weeks=repeat_every)
    elif period == 'month':
        month = today.month + repeat_every
        year = today.year + (month - 1) // 12
        month = ((month - 1) % 12) + 1
        day = min(today.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)
    elif period == 'year':
        return date(today.year + repeat_every, today.month, today.day)
    
    return today + timedelta(days=30)


def get_invoice_by_id_service(invoice_id: int, db: Session):
    """Get invoice with all details by ID"""
    try:
        # Get invoice master with building/level/unit names
        invoice_query = """
            SELECT 
                im.id, im.invoice_number, im.resident_id, im.building_id, im.level_id, im.unit_id,
                DATE_FORMAT(im.invoice_date, '%Y-%m-%d') as invoice_date,
                DATE_FORMAT(im.due_date, '%Y-%m-%d') as due_date,
                im.sub_total, im.extra_amount, im.gst, im.total_amount, im.discount,
                im.terms_and_conditions, im.status, im.payment_status, im.mark_as_recurring,
                im.parent_invoice_id, im.created_by,
                DATE_FORMAT(im.created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                DATE_FORMAT(im.updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at,
                pb.building_name,
                bl.level as level_name,
                bu.unit_no as unit_number
            FROM invoice_master im
            LEFT JOIN property_building pb ON im.building_id = pb.id
            LEFT JOIN building_level bl ON im.level_id = bl.id
            LEFT JOIN building_units bu ON im.unit_id = bu.id
            WHERE im.id = :id
        """
        
        invoice = db.execute(text(invoice_query), {"id": invoice_id}).mappings().first()
        
        if not invoice:
            return None
        
        invoice_dict = dict(invoice)
        
        # Get vehicle items
        vehicle_items_query = """
            SELECT id, invoice_id, vehicle_id, iu_number, vehicle_number, vehicle_type_id,
                   sub_total, extra_amount, gst, total_amount, discount, description, created_by,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM invoice_vehicle_items 
            WHERE invoice_id = :invoice_id
        """
        
        vehicle_items = db.execute(text(vehicle_items_query), {"invoice_id": invoice_id}).mappings().all()
        invoice_dict['vehicle_items'] = [dict(item) for item in vehicle_items]
        
        # Get recurring settings - use backticks for reserved keywords
        recurring_query = """
            SELECT id, invoice_id, repeat_every, period, month_option, day_of_month,
                   weekday, week_number, selected_days, end_type,
                   DATE_FORMAT(end_date, '%Y-%m-%d') as end_date,
                   after_occurrences, `year_month`, `year_day`,
                   DATE_FORMAT(next_invoice_date, '%Y-%m-%d') as next_invoice_date,
                   DATE_FORMAT(last_generated_date, '%Y-%m-%d') as last_generated_date,
                   total_generated, is_active,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM recurring_invoice_settings 
            WHERE invoice_id = :invoice_id
        """
        
        recurring = db.execute(text(recurring_query), {"invoice_id": invoice_id}).mappings().first()
        if recurring:
            invoice_dict['recurring_settings'] = dict(recurring)
        else:
            invoice_dict['recurring_settings'] = None
        
        # Get recurring residents
        recurring_residents_query = """
            SELECT id, resident_id, invoice_id, building_id, level_id, unit_id, created_by,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM invoice_recurring_residents 
            WHERE invoice_id = :invoice_id
        """
        
        recurring_residents = db.execute(text(recurring_residents_query), {"invoice_id": invoice_id}).mappings().first()
        if recurring_residents:
            invoice_dict['recurring_resident'] = dict(recurring_residents)
        else:
            invoice_dict['recurring_resident'] = None
        
        return invoice_dict
        
    except Exception as e:
        write_to_server_log(f"Error in get_invoice_by_id_service: {str(e)}")
        import traceback
        write_to_server_log(f"Traceback: {traceback.format_exc()}")
        raise Exception(f"Failed to get invoice: {str(e)}")


def list_invoices_service(
    db: Session,
    invoice_id: Optional[int] = None,
    resident_id: Optional[int] = None,
    building_id: Optional[int] = None,
    status: Optional[str] = None,
    payment_status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    pagination: bool = True,
    page: int = 1,
    per_page: Optional[int] = 10
):
    """List invoices with filters and pagination, including vehicle items, vehicle details, and recurring settings"""
    try:
        # Step 1: Build the base filter conditions for invoice IDs
        filter_conditions = "WHERE 1=1"
        params = {}
        
        # Apply filters
        if invoice_id is not None:
            filter_conditions += " AND im.id = :invoice_id"
            params["invoice_id"] = invoice_id
        
        if resident_id is not None:
            filter_conditions += " AND im.resident_id = :resident_id"
            params["resident_id"] = resident_id
        
        if building_id is not None:
            filter_conditions += " AND im.building_id = :building_id"
            params["building_id"] = building_id
        
        if status:
            filter_conditions += " AND LOWER(im.status) = :status"
            params["status"] = status.lower()
        
        if payment_status:
            filter_conditions += " AND LOWER(im.payment_status) = :payment_status"
            params["payment_status"] = payment_status.lower()
        
        if from_date:
            filter_conditions += " AND DATE(im.invoice_date) >= :from_date"
            params["from_date"] = from_date
        
        if to_date:
            filter_conditions += " AND DATE(im.invoice_date) <= :to_date"
            params["to_date"] = to_date
        
        # Step 2: Get total count of distinct invoices (only needed for pagination)
        total_count = 0
        if pagination:
            count_query = f"""
                SELECT COUNT(DISTINCT im.id) as total 
                FROM invoice_master im
                {filter_conditions}
            """
            count_result = db.execute(text(count_query), params).first()
            total_count = count_result[0] if count_result else 0
        
        # Step 3: Get invoice IDs (with or without pagination)
        invoice_ids_query = f"""
            SELECT DISTINCT im.id
            FROM invoice_master im
            {filter_conditions}
            ORDER BY im.id DESC
        """
        
        # Apply pagination only if enabled
        if pagination and per_page and invoice_id is None:
            invoice_ids_query += " LIMIT :limit OFFSET :offset"
            offset = (page - 1) * per_page
            params["limit"] = per_page
            params["offset"] = offset
        
        # Execute to get invoice IDs
        invoice_ids_result = db.execute(text(invoice_ids_query), params).fetchall()
        invoice_ids = [row[0] for row in invoice_ids_result]
        
        # If no invoices found, return empty result
        if not invoice_ids:
            if pagination:
                return {
                    "invoices": [],
                    "total": total_count,
                    "total_pages": 0 if total_count == 0 else (total_count + per_page - 1) // per_page if per_page else 1
                }
            else:
                return {
                    "invoices": [],
                    "total": 0,
                    "total_pages": 1
                }
        
        # Step 4: Fetch full details for these specific invoice IDs
        # Create placeholders for IN clause
        placeholders = ','.join([':id{}'.format(i) for i in range(len(invoice_ids))])
        id_params = {'id{}'.format(i): invoice_id_val for i, invoice_id_val in enumerate(invoice_ids)}
        
        base_query = f"""
            SELECT 
                im.id, im.invoice_number, im.resident_id, im.building_id, im.level_id, im.paid_at, im.unit_id,
                DATE_FORMAT(im.invoice_date, '%Y-%m-%d') as invoice_date,
                DATE_FORMAT(im.due_date, '%Y-%m-%d') as due_date,
                im.sub_total, im.extra_amount, im.gst, im.total_amount, im.discount,
                im.status, im.payment_status, im.mark_as_recurring,
                DATE_FORMAT(im.created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                pb.building_name,
                bl.level as level_name,
                bu.unit_no as unit_number,
                
                /* Resident Details from user_personal_details */
                upd_resident.first_name as resident_first_name,
                upd_resident.last_name as resident_last_name,
                upd_resident.email as resident_email,
                upd_resident.phone as resident_phone,
                
                /* Vehicle Items Details */
                ivi.id as vehicle_item_id,
                ivi.vehicle_id,
                ivi.iu_number as vehicle_iu_number,
                ivi.vehicle_number,
                ivi.vehicle_type_id,
                ivi.sub_total as vehicle_sub_total,
                ivi.extra_amount as vehicle_extra_amount,
                ivi.gst as vehicle_gst,
                ivi.total_amount as vehicle_total_amount,
                ivi.discount as vehicle_discount,
                ivi.description as vehicle_description,
                
                /* License Plate Access Details (Vehicle Master Data) */
                lpa.LicensePlate as license_plate_number,
                lpa.listType as vehicle_list_type,
                lpa.vehicle_type as vehicle_type_name,
                lpa.source as vehicle_source,
                lpa.effectiveTime as vehicle_effective_time,
                lpa.anpr_device_activities as anpr_device_activities,
                lpa.resident_id as vehicle_owner_resident_id,
                
                /* Get owner name from user_personal_details */
                CONCAT(upd.first_name, ' ', upd.last_name) as vehicle_owner_name,
                
                /* Recurring Invoice Settings */
                ris.id as recurring_setting_id,
                ris.repeat_every,
                ris.period,
                ris.month_option,
                ris.day_of_month,
                ris.weekday,
                ris.week_number,
                ris.selected_days,
                ris.year_month,
                ris.year_day,
                ris.end_type,
                DATE_FORMAT(ris.end_date, '%Y-%m-%d') as end_date,
                ris.after_occurrences,
                DATE_FORMAT(ris.next_invoice_date, '%Y-%m-%d') as next_invoice_date,
                DATE_FORMAT(ris.last_generated_date, '%Y-%m-%d') as last_generated_date,
                ris.total_generated,
                ris.is_active
                
            FROM invoice_master im
            LEFT JOIN user_personal_details upd_resident ON im.resident_id = upd_resident.id
            LEFT JOIN invoice_vehicle_items ivi ON im.id = ivi.invoice_id
            LEFT JOIN license_plate_access lpa ON ivi.vehicle_number = lpa.LicensePlate
            LEFT JOIN user_personal_details upd ON lpa.resident_id = upd.id
            LEFT JOIN recurring_invoice_settings ris ON im.id = ris.invoice_id
            LEFT JOIN property_building pb ON im.building_id = pb.id
            LEFT JOIN building_level bl ON im.level_id = bl.id
            LEFT JOIN building_units bu ON im.unit_id = bu.id
            WHERE im.id IN ({placeholders})
            ORDER BY im.id DESC
        """
        
        # Merge parameters (filter params + ID params)
        all_params = {**params, **id_params}
        
        # Execute query
        result = db.execute(text(base_query), all_params).mappings().all()
        
        # Helper function to parse JSON extra_amount
        def parse_extra_amount(extra_amount_value):
            """Parse extra_amount from JSON string or return default value"""
            if extra_amount_value is None:
                return []
            
            # If it's already a list, return it
            if isinstance(extra_amount_value, list):
                return extra_amount_value
            
            # If it's a string, try to parse as JSON
            if isinstance(extra_amount_value, str):
                try:
                    parsed = json.loads(extra_amount_value)
                    if isinstance(parsed, list):
                        return parsed
                    else:
                        return [{"reason": "Extra charges", "amount": float(parsed)}] if parsed else []
                except:
                    # If it's a simple number string
                    try:
                        num_val = float(extra_amount_value)
                        return [{"reason": "Extra charges", "amount": num_val}] if num_val else []
                    except:
                        return []
            
            # If it's a number, convert to list format
            if isinstance(extra_amount_value, (int, float)):
                return [{"reason": "Extra charges", "amount": float(extra_amount_value)}] if extra_amount_value else []
            
            return []
        
        # Helper function to get total from extra_amount list
        def get_extra_amount_total(extra_amount_list):
            """Calculate total from extra_amount list"""
            if not extra_amount_list:
                return 0
            total = 0
            for item in extra_amount_list:
                if isinstance(item, dict):
                    total += float(item.get('amount', 0))
                elif isinstance(item, (int, float)):
                    total += float(item)
            return total
        
        # Group results by invoice
        invoices_dict = {}
        for row in result:
            invoice_id_val = row['id']
            
            if invoice_id_val not in invoices_dict:
                # Parse extra_amount for invoice level
                invoice_extra_amount = parse_extra_amount(row['extra_amount'])
                
                # Build resident object
                resident_obj = {
                    "id": row['resident_id'],
                    "first_name": row['resident_first_name'],
                    "last_name": row['resident_last_name'],
                    "email": row['resident_email'],
                    "phone": row['resident_phone']
                }
                
                invoices_dict[invoice_id_val] = {
                    "id": row['id'],
                    "invoice_number": row['invoice_number'],
                    "resident_id": row['resident_id'],
                    "resident": resident_obj,
                    "building_id": row['building_id'],
                    "level_id": row['level_id'],
                    "paid_at": row['paid_at'],
                    "unit_id": row['unit_id'],
                    "building_name": row['building_name'],
                    "level_name": row['level_name'],
                    "unit_number": row['unit_number'],
                    "invoice_date": row['invoice_date'],
                    "due_date": row['due_date'],
                    "sub_total": float(row['sub_total']) if row['sub_total'] else 0,
                    "extra_amount": invoice_extra_amount,
                    "extra_amount_total": get_extra_amount_total(invoice_extra_amount),
                    "gst": float(row['gst']) if row['gst'] else 0,
                    "total_amount": float(row['total_amount']) if row['total_amount'] else 0,
                    "discount": float(row['discount']) if row['discount'] else 0,
                    "status": row['status'].upper() if row['status'] else None,
                    "payment_status": row['payment_status'].upper() if row['payment_status'] else None,
                    "mark_as_recurring": row['mark_as_recurring'] == 1 or row['mark_as_recurring'] == True,
                    "created_at": row['created_at'],
                    "vehicle_items": [],
                    "recurring_setting": None
                }
            
            # Add vehicle item if exists (with duplicate prevention)
            if row['vehicle_item_id']:
                # Check if this vehicle item already exists in the invoice's vehicle_items
                existing_vehicle_ids = [v['id'] for v in invoices_dict[invoice_id_val]['vehicle_items']]
                
                if row['vehicle_item_id'] not in existing_vehicle_ids:
                    # Parse extra_amount for vehicle item
                    vehicle_extra_amount = parse_extra_amount(row['vehicle_extra_amount'])
                    
                    # Build vehicle details with owner information
                    vehicle_details = None
                    if row['license_plate_number']:
                        vehicle_details = {
                            "license_plate_number": row['license_plate_number'],
                            "list_type": row['vehicle_list_type'],
                            "vehicle_type_name": row['vehicle_type_name'],
                            "source": row['vehicle_source'],
                            "effective_time": row['vehicle_effective_time'],
                            "anpr_device_activities": row['anpr_device_activities'],
                            "owner_resident_id": row['vehicle_owner_resident_id'],
                            "owner_name": row['vehicle_owner_name'] or 'Unknown'
                        }
                    
                    vehicle_item = {
                        "id": row['vehicle_item_id'],
                        "vehicle_id": row['vehicle_id'],
                        "iu_number": row['vehicle_iu_number'],
                        "vehicle_number": row['vehicle_number'],
                        "vehicle_type_id": row['vehicle_type_id'],
                        "sub_total": float(row['vehicle_sub_total']) if row['vehicle_sub_total'] else 0,
                        "extra_amount": vehicle_extra_amount,
                        "extra_amount_total": get_extra_amount_total(vehicle_extra_amount),
                        "gst": float(row['vehicle_gst']) if row['vehicle_gst'] else 0,
                        "total_amount": float(row['vehicle_total_amount']) if row['vehicle_total_amount'] else 0,
                        "discount": float(row['vehicle_discount']) if row['vehicle_discount'] else 0,
                        "description": row['vehicle_description'],
                        "vehicle_details": vehicle_details
                    }
                    invoices_dict[invoice_id_val]['vehicle_items'].append(vehicle_item)
            
            # Add recurring setting if exists (only once per invoice)
            if row['recurring_setting_id'] and invoices_dict[invoice_id_val]['recurring_setting'] is None:
                # Parse selected_days JSON if exists
                selected_days_value = row['selected_days']
                if selected_days_value:
                    try:
                        selected_days_value = json.loads(selected_days_value)
                    except:
                        selected_days_value = None
                
                recurring_setting = {
                    "id": row['recurring_setting_id'],
                    "repeat_every": row['repeat_every'],
                    "period": row['period'],
                    "month_option": row['month_option'],
                    "day_of_month": row['day_of_month'],
                    "weekday": row['weekday'],
                    "week_number": row['week_number'],
                    "selected_days": selected_days_value,
                    "year_month": row['year_month'],
                    "year_day": row['year_day'],
                    "end_type": row['end_type'],
                    "end_date": row['end_date'],
                    "after_occurrences": row['after_occurrences'],
                    "next_invoice_date": row['next_invoice_date'],
                    "last_generated_date": row['last_generated_date'],
                    "total_generated": row['total_generated'] or 0,
                    "is_active": row['is_active'] == 1 or row['is_active'] == True
                }
                invoices_dict[invoice_id_val]['recurring_setting'] = recurring_setting
        
        invoices_list = list(invoices_dict.values())
        
        # If filtering by single invoice_id, return single object instead of list
        if invoice_id is not None and len(invoices_list) == 1:
            return invoices_list[0]
        
        # Calculate total pages for pagination (only if pagination is enabled)
        if pagination and per_page and invoice_id is None:
            total_pages = (total_count + per_page - 1) // per_page if per_page > 0 else 0
        else:
            total_pages = 1
        
        return {
            "invoices": invoices_list,
            "total": total_count if pagination else len(invoices_list),
            "total_pages": total_pages
        }
        
    except Exception as e:
        write_to_server_log(f"Error in list_invoices_service: {str(e)}")
        import traceback
        write_to_server_log(f"Traceback: {traceback.format_exc()}")
        raise Exception(f"Failed to list invoices: {str(e)}")


def _format_invoices_for_export(invoices: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for invoice in invoices:
        resident = invoice.get("resident") or {}
        resident_name = f"{resident.get('first_name', '')} {resident.get('last_name', '')}".strip() or "-"
        rows.append({
            "invoice_number": invoice.get("invoice_number") or "-",
            "resident_name": resident_name,
            "building_name": invoice.get("building_name") or "-",
            "level_name": invoice.get("level_name") or "-",
            "unit_number": invoice.get("unit_number") or "-",
            "invoice_date": invoice.get("invoice_date") or "-",
            "due_date": invoice.get("due_date") or "-",
            "total_amount": invoice.get("total_amount") if invoice.get("total_amount") is not None else "-",
            "status": invoice.get("status") or "-",
            "payment_status": invoice.get("payment_status") or "-",
            "created_at": invoice.get("created_at") or "-",
        })
    return rows


def generate_invoices_excel_response(data: List[Dict[str, Any]], filename: str = "invoices.xlsx"):
    try:
        rows = _format_invoices_for_export(data)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "S.No",
            "Invoice Number",
            "Resident Name",
            "Building",
            "Level",
            "Unit",
            "Invoice Date",
            "Due Date",
            "Total Amount",
            "Status",
            "Payment Status",
            "Created At",
        ]
        field_mapping = {
            "Invoice Number": "invoice_number",
            "Resident Name": "resident_name",
            "Building": "building_name",
            "Level": "level_name",
            "Unit": "unit_number",
            "Invoice Date": "invoice_date",
            "Due Date": "due_date",
            "Total Amount": "total_amount",
            "Status": "status",
            "Payment Status": "payment_status",
            "Created At": "created_at",
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for row_index, record in enumerate(rows):
            row_num = row_index + 2
            for col_num, header in enumerate(headers, 1):
                if header == "S.No":
                    value = row_index + 1
                else:
                    value = record.get(field_mapping.get(header, ""), "-")
                    if value is None or value == "":
                        value = "-"
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = cell_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(rows) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        ws.freeze_panes = "A2"
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")


def generate_invoices_pdf_response(data: List[Dict[str, Any]], filename: str = "invoices.pdf"):
    try:
        rows = _format_invoices_for_export(data)
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        elements = []
        styles = getSampleStyleSheet()

        logo_url = get_logo_url_prod()
        try:
            logo = Image(logo_url)
            logo.drawHeight = 1.0 * inch
            logo.drawWidth = 2.0 * inch
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

        title_style = styles['Heading1']
        title_style.alignment = 1
        title_style.fontSize = 16
        title_style.spaceAfter = 0.2 * inch
        elements.append(Paragraph("Invoices Report", title_style))

        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            alignment=1,
            fontSize=10,
            textColor=colors.gray,
            spaceAfter=0.2 * inch
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.2 * inch))

        headers = [
            "S.No", "Invoice No", "Resident", "Building", "Level", "Unit",
            "Inv Date", "Due Date", "Total", "Status", "Payment"
        ]

        page_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
        col_widths = [
            page_width * 0.05,
            page_width * 0.12,
            page_width * 0.14,
            page_width * 0.12,
            page_width * 0.08,
            page_width * 0.08,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.08,
            page_width * 0.07,
            page_width * 0.06,
        ]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ])

        max_rows_per_page = 12
        for start in range(0, len(rows), max_rows_per_page):
            chunk = rows[start:start + max_rows_per_page]
            table_data = [headers]
            for idx, record in enumerate(chunk, start=start + 1):
                table_data.append([
                    str(idx),
                    str(record.get("invoice_number", "-")),
                    str(record.get("resident_name", "-")),
                    str(record.get("building_name", "-")),
                    str(record.get("level_name", "-")),
                    str(record.get("unit_number", "-")),
                    str(record.get("invoice_date", "-")),
                    str(record.get("due_date", "-")),
                    str(record.get("total_amount", "-")),
                    str(record.get("status", "-")),
                    str(record.get("payment_status", "-")),
                ])

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(table_style)
            elements.append(table)

            if start + max_rows_per_page < len(rows):
                elements.append(PageBreak())

        elements.append(Spacer(1, 0.2 * inch))
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.darkblue
        )
        elements.append(Paragraph(f"Total Invoices: {len(rows)}", summary_style))

        doc.build(elements)
        pdf_buffer.seek(0)

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


def update_invoice_status_service(invoice_id: int, status: str, db: Session):
    """Update invoice status"""
    try:
        existing = get_invoice_by_id_service(invoice_id, db)
        if not existing:
            return None
        
        db.execute(
            text("""
                UPDATE invoice_master 
                SET status = :status, updated_at = NOW()
                WHERE id = :id
            """),
            {"id": invoice_id, "status": status}
        )
        db.commit()
        
        return get_invoice_by_id_service(invoice_id, db)
        
    except Exception as e:
        db.rollback()
        error_msg = str(e)
        write_to_server_log(f"Error in update_invoice_status_service: {error_msg}")
        raise Exception(f"Failed to update invoice status: {error_msg}")


def delete_invoice_service(invoice_id: int, db: Session):
    """Delete invoice"""
    try:
        existing = get_invoice_by_id_service(invoice_id, db)
        if not existing:
            return None
        
        db.execute(text("DELETE FROM invoice_master WHERE id = :id"), {"id": invoice_id})
        db.commit()
        
        return {"id": invoice_id, "deleted": True}
        
    except Exception as e:
        db.rollback()
        error_msg = str(e)
        write_to_server_log(f"Error in delete_invoice_service: {error_msg}")
        raise Exception(f"Failed to delete invoice: {error_msg}")


def get_recurring_residents_by_invoice_id(invoice_id: int, db: Session):
    """Get recurring resident details by invoice ID"""
    try:
        query = """
            SELECT id, resident_id, invoice_id, building_id, level_id, unit_id, created_by,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM invoice_recurring_residents 
            WHERE invoice_id = :invoice_id
        """
        
        result = db.execute(text(query), {"invoice_id": invoice_id}).mappings().first()
        return dict(result) if result else None
        
    except Exception as e:
        write_to_server_log(f"Error in get_recurring_residents_by_invoice_id: {str(e)}")
        raise Exception(f"Failed to get recurring residents: {str(e)}")


def list_all_recurring_residents(db: Session, pagination: bool = True, page: int = 1, per_page: int = 10):
    """List all recurring residents"""
    try:
        base_query = """
            SELECT id, resident_id, invoice_id, building_id, level_id, unit_id, created_by,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM invoice_recurring_residents
            ORDER BY id DESC
        """
        
        count_query = "SELECT COUNT(*) as total FROM invoice_recurring_residents"
        params = {}
        
        # Get total count
        count_result = db.execute(text(count_query), params).first()
        total_count = count_result[0] if count_result else 0
        
        # Add pagination
        if pagination and per_page:
            base_query += " LIMIT :limit OFFSET :offset"
            offset = (page - 1) * per_page
            params["limit"] = per_page
            params["offset"] = offset
        
        result = db.execute(text(base_query), params).mappings().all()
        recurring_residents_list = [dict(item) for item in result]
        
        if pagination and per_page:
            total_pages = (total_count + per_page - 1) // per_page if per_page > 0 else 0
        else:
            total_pages = 1
        
        return {
            "recurring_residents": recurring_residents_list,
            "total": total_count,
            "total_pages": total_pages
        }
        
    except Exception as e:
        write_to_server_log(f"Error in list_all_recurring_residents: {str(e)}")
        raise Exception(f"Failed to list recurring residents: {str(e)}")
    

def get_invoice_by_id_with_payment_service(invoice_id: int, db: Session):
    """Get invoice with payment details by ID"""
    try:
        # Get invoice master
        invoice_query = """
            SELECT id, invoice_number, resident_id, building_id, level_id, unit_id,
                   DATE_FORMAT(invoice_date, '%Y-%m-%d') as invoice_date,
                   DATE_FORMAT(due_date, '%Y-%m-%d') as due_date,
                   sub_total, extra_amount, gst, total_amount, discount,
                   terms_and_conditions, status, payment_status, mark_as_recurring,
                   parent_invoice_id, paid_at,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                   DATE_FORMAT(updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at
            FROM invoice_master 
            WHERE id = :id
        """
        
        invoice = db.execute(text(invoice_query), {"id": invoice_id}).mappings().first()
        
        if not invoice:
            return None
        
        invoice_dict = dict(invoice)
        
        # Get vehicle items
        vehicle_items_query = """
            SELECT id, invoice_id, vehicle_id, iu_number, vehicle_number, vehicle_type_id,
                   sub_total, extra_amount, gst, total_amount, discount, description,
                   DATE_FORMAT(created_at, '%Y-%m-%dT%H:%i:%s') as created_at
            FROM invoice_vehicle_items 
            WHERE invoice_id = :invoice_id
        """
        
        vehicle_items = db.execute(text(vehicle_items_query), {"invoice_id": invoice_id}).mappings().all()
        invoice_dict['vehicle_items'] = [dict(item) for item in vehicle_items]
        
        # Get payment details
        payment_query = """
            SELECT id, payment_request_id, reference_number, amount, 
                   status as payment_record_status, payment_url, created_at as payment_created_at
            FROM invoice_payments 
            WHERE invoice_id = :invoice_id
            ORDER BY id DESC
            LIMIT 1
        """
        
        payment = db.execute(text(payment_query), {"invoice_id": invoice_id}).mappings().first()
        invoice_dict['payment_details'] = dict(payment) if payment else None
        
        return invoice_dict
        
    except Exception as e:
        write_to_server_log(f"Error in get_invoice_by_id_with_payment_service: {str(e)}")
        raise Exception(f"Failed to get invoice: {str(e)}")
    
def get_invoice_by_id_with_payment_service(invoice_id: int, db: Session):
    """Get invoice with payment details and resident info by ID"""
    try:
        # Get invoice master with resident details - FIXED column names
        invoice_query = """
            SELECT 
                im.id, im.invoice_number, im.resident_id, im.building_id, im.level_id, im.unit_id,
                DATE_FORMAT(im.invoice_date, '%Y-%m-%d') as invoice_date,
                DATE_FORMAT(im.due_date, '%Y-%m-%d') as due_date,
                im.sub_total, im.extra_amount, im.gst, im.total_amount, im.discount,
                im.terms_and_conditions, im.status, im.payment_status, im.mark_as_recurring,
                im.parent_invoice_id, im.paid_at,
                DATE_FORMAT(im.created_at, '%Y-%m-%dT%H:%i:%s') as created_at,
                DATE_FORMAT(im.updated_at, '%Y-%m-%dT%H:%i:%s') as updated_at,
                up.first_name, up.last_name, up.email, up.phone,
                pb.building_name,  -- Changed from b.name to pb.building_name
                bl.level as level_name,  -- Changed from bl.name to bl.level
                bu.unit_no as unit_number  -- Changed from bu.unit_number to bu.unit_no
            FROM invoice_master im
            LEFT JOIN user_personal_details up ON im.resident_id = up.id
            LEFT JOIN property_building pb ON im.building_id = pb.id  -- Changed alias
            LEFT JOIN building_level bl ON im.level_id = bl.id
            LEFT JOIN building_units bu ON im.unit_id = bu.id
            WHERE im.id = :id
        """
        
        invoice = db.execute(text(invoice_query), {"id": invoice_id}).mappings().first()
        
        if not invoice:
            return None
        
        invoice_dict = dict(invoice)
        
        # Add resident object
        invoice_dict['resident'] = {
            'first_name': invoice.get('first_name'),
            'last_name': invoice.get('last_name'),
            'email': invoice.get('email'),
            'phone': invoice.get('phone')
        }
        
        # Remove flat fields
        for field in ['first_name', 'last_name', 'email', 'phone']:
            if field in invoice_dict:
                del invoice_dict[field]
        
        # Get vehicle items with vehicle details
        vehicle_items_query = """
            SELECT 
                ivi.id, ivi.invoice_id, ivi.vehicle_id, ivi.iu_number, ivi.vehicle_number, 
                ivi.vehicle_type_id, ivi.sub_total, ivi.extra_amount, ivi.gst, 
                ivi.total_amount, ivi.discount, ivi.description,
                vt.title as vehicle_type_name,
                DATE_FORMAT(ivi.created_at, '%Y-%m-%dT%H:%i:%s') as created_at
            FROM invoice_vehicle_items ivi
            LEFT JOIN vehicle_type vt ON ivi.vehicle_type_id = vt.id
            WHERE ivi.invoice_id = :invoice_id
        """
        
        vehicle_items = db.execute(text(vehicle_items_query), {"invoice_id": invoice_id}).mappings().all()
        invoice_dict['vehicle_items'] = [dict(item) for item in vehicle_items]
        
        # Get payment details
        payment_query = """
            SELECT id, payment_request_id, reference_number, amount, 
                   status as payment_record_status, payment_url, created_at as payment_created_at
            FROM invoice_payments 
            WHERE invoice_id = :invoice_id
            ORDER BY id DESC
            LIMIT 1
        """
        
        payment = db.execute(text(payment_query), {"invoice_id": invoice_id}).mappings().first()
        invoice_dict['payment_details'] = dict(payment) if payment else None
        
        return invoice_dict
        
    except Exception as e:
        write_to_server_log(f"Error in get_invoice_by_id_with_payment_service: {str(e)}")
        import traceback
        write_to_server_log(f"Traceback: {traceback.format_exc()}")
        raise Exception(f"Failed to get invoice: {str(e)}")