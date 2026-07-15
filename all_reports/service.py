from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
import os,re
import uuid as uuid_lib
from typing import List
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
from fastapi.responses import StreamingResponse
from utils.common_function import format_date, get_logo_url_prod
from typing import Optional, List, Dict, Any
import json


role_map = {
    1: "BMO ADMIN",
    2: "Manager",
    3: "Security"
}

def format_activity_log(record: Dict[str, Any]) -> Dict[str, Any]:
    """Format activity log record for response"""
    
    actor_email = "system@example.com"
    created_by = "Unknown"
    role = "Unknown"
    
    new_data = record.get('new_data', {})
    
    if isinstance(new_data, str):
        try:
            new_data = json.loads(new_data)
        except:
            new_data = {}
    
  
    if 'creator_info' in new_data:
        creator_info = new_data['creator_info']
        actor_email = creator_info.get('creator_email', 'system@example.com')
        created_by = creator_info.get('creator_name', 'Unknown')
        
        role_id = creator_info.get('creator_company_id') or creator_info.get('creator_role_id')
        if role_id:
            role_map = {
                1: "Super Admin",
                2: "Security",
                3: "Manager",
                4: "Conceirge"
            }
            role = role_map.get(role_id, "Unknown")
    
    elif 'updator_info' in new_data:
        updator_info = new_data['updator_info']
        actor_email = updator_info.get('updator_email', 'system@example.com')
        created_by = updator_info.get('updator_name', 'Unknown')
        
        role_id = updator_info.get('updater_company_id') or updator_info.get('updater_role_id')
        if role_id:
            role_map = {
                1: "Super Admin",
                2: "Security",
                3: "Manager",
                4: "Conceirge"
            }
            role = role_map.get(role_id, "Unknown")
    
    if role == "Unknown" and record.get('role_id'):
        role_map = {
            1: "Super Admin",
            2: "Security",
            3: "Manager",
            4: "Conceirge"
        }
        role = role_map.get(record.get('role_id'), "Unknown")
    
    if role == "Unknown" and actor_email and actor_email != "system@example.com":
        if actor_email.lower() == "bmoadmin@yopmail.com":
            role = "BMO ADMIN"
        elif "admin" in actor_email.lower():
            role = "Super Admin"
        elif "manager" in actor_email.lower():
            role = "Manager"
        elif "security" in actor_email.lower():
            role = "Security"
    
    if created_by == "Unknown" and record.get('created_by'):
        created_by = record.get('created_by')
    
    created_at = record.get('created_at')
    if created_at:
        if isinstance(created_at, datetime):
            created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(created_at, str):
            try:
                dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                created_at = dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
    
    return {
        "id": record.get('id'),
        "module_name": record.get('module_name', 'N/A'),
        "event_type": record.get('action', 'N/A'),
        "description": record.get('description', ''),
        "role": role,
        "company": None,
        "actor_email": actor_email,
        "created_at": created_at,
        "created_by": created_by,
        "new_data": new_data if isinstance(new_data, dict) else {},
    }


def generate_activity_logs_excel_response(data: List[Dict[str, Any]], filename: str = "activity_logs.xlsx"):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activity Logs"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ["S.No", "Module Name", "Event Type", "Description", "Role", "Log Activity"]
        field_mapping = {
            "Module Name": "module_name",
            "Event Type": "event_type",
            "Description": "description",
            "Role": "role",
            "Log Activity": "created_at",
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for row_index, record in enumerate(data):
            row_num = row_index + 2
            for col_num, header in enumerate(headers, 1):
                if header == "S.No":
                    value = row_index + 1
                else:
                    field_name = field_mapping.get(header)
                    value = record.get(field_name, "") if field_name else ""
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    if value is None or value == "":
                        value = "-"

                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = cell_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(data) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")


def generate_activity_logs_pdf_response(data: List[Dict[str, Any]], filename: str = "activity_logs.pdf"):
    try:
        pdf_buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        elements = []
        styles = getSampleStyleSheet()

        logo_url = get_logo_url_prod()
        try:
            logo = Image(logo_url)
            logo.drawHeight = 1.0 * inch
            logo.drawWidth = 2.0 * inch
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

        title_style = styles['Heading1']
        title_style.alignment = 1
        title_style.fontSize = 16
        title_style.spaceAfter = 0.2 * inch
        elements.append(Paragraph("Activity Logs Report", title_style))

        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            alignment=1,
            fontSize=10,
            textColor=colors.gray,
            spaceAfter=0.2 * inch
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.2 * inch))

        headers = ["S.No", "Module Name", "Event Type", "Description", "Role", "Log Activity"]

        page_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
        col_widths = [
            page_width * 0.07,
            page_width * 0.18,
            page_width * 0.14,
            page_width * 0.38,
            page_width * 0.11,
            page_width * 0.12,
        ]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ])

        max_rows_per_page = 12
        total_records = len(data)

        for start in range(0, total_records, max_rows_per_page):
            end = start + max_rows_per_page
            chunk = data[start:end]

            table_data = [headers]
            for idx, record in enumerate(chunk, start=start + 1):
                module_name = record.get('module_name', '') or "-"
                event_type = record.get('event_type', '') or "-"
                description = record.get('description', '') or "-"
                role = record.get('role', '') or "-"
                log_activity = record.get('created_at', '') or "-"

                row = [str(idx), module_name, event_type, description, role, str(log_activity)]
                table_data.append(row)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(table_style)
            elements.append(table)

            if end < total_records:
                elements.append(PageBreak())

        elements.append(Spacer(1, 0.2 * inch))
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.darkblue
        )
        elements.append(Paragraph(f"Total Activity Logs: {len(data)}", summary_style))

        doc.build(elements)
        pdf_buffer.seek(0)

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")
    
def list_visitors_service(
    db: Session,
    id: Optional[int] = None,
    visitor_id: Optional[str] = None,
    visitor_type: Optional[str] = None,
    search: Optional[str] = None,
    building_id: Optional[int] = None,
    level_id: Optional[int] = None,
    unit_id: Optional[int] = None,
    is_qr_valid: Optional[bool] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    record_count: int = 10,
    download: bool = False,
):
    try:
        if visitor_type and str(visitor_type).strip().lower() in ["null", "none", "all", ""]:
            visitor_type = None

        offset = (page - 1) * record_count
        params = {}

        query = """
            SELECT * FROM (
                SELECT 
                    av.id,
                    av.name,
                    av.email,
                    av.adoc_visitor_id AS visitor_id,
                    av.phone,
                    av.purpose_visit,
                    av.valid,
                    av.qr_token,
                    av.created_at,
                    av.updated_at,
                    'adoc' AS visitor_type_raw,
                    'Ad-hoc' AS visitor_type_display,
                    av.building_id,
                    av.level_id,
                    av.unit_id,
                    pb.building_name,
                    bl.level AS level_name,
                    bu.unit_no AS unit_name
                FROM adoc_visitor av
                LEFT JOIN property_building pb ON av.building_id = pb.id
                LEFT JOIN building_level bl ON av.level_id = bl.id
                LEFT JOIN building_units bu ON av.unit_id = bu.id

                UNION ALL

                SELECT 
                    iv.id,
                    iv.name,
                    NULL AS email,
                    iv.visitor_id,
                    NULL AS phone,
                    iv.purpose_visit,
                    iv.valid,
                    iv.qr_token,
                    iv.created_at,
                    iv.updated_at,
                    'invite' AS visitor_type_raw,
                    'Invite' AS visitor_type_display,
                    ua.building_id,
                    ua.level_id,
                    ua.unit_id,
                    pb.building_name,
                    bl.level AS level_name,
                    bu.unit_no AS unit_name
                FROM invite_visitor iv
                LEFT JOIN user_access_details ua ON iv.user_id = ua.user_id
                LEFT JOIN property_building pb ON ua.building_id = pb.id
                LEFT JOIN building_level bl ON ua.level_id = bl.id
                LEFT JOIN building_units bu ON ua.unit_id = bu.id
            ) AS visitors
            WHERE 1=1
        """

        count_query = """
            SELECT COUNT(*) FROM (
                SELECT 
                    av.id,
                    av.name,
                    av.email,
                    av.adoc_visitor_id AS visitor_id,
                    av.phone,
                    av.purpose_visit,
                    av.valid,
                    av.qr_token,
                    av.created_at,
                    av.updated_at,
                    'adoc' AS visitor_type_raw,
                    'Ad-hoc' AS visitor_type_display,
                    av.building_id,
                    av.level_id,
                    av.unit_id,
                    pb.building_name,
                    bl.level AS level_name,
                    bu.unit_no AS unit_name
                FROM adoc_visitor av
                LEFT JOIN property_building pb ON av.building_id = pb.id
                LEFT JOIN building_level bl ON av.level_id = bl.id
                LEFT JOIN building_units bu ON av.unit_id = bu.id

                UNION ALL

                SELECT 
                    iv.id,
                    iv.name,
                    NULL AS email,
                    iv.visitor_id,
                    NULL AS phone,
                    iv.purpose_visit,
                    iv.valid,
                    iv.qr_token,
                    iv.created_at,
                    iv.updated_at,
                    'invite' AS visitor_type_raw,
                    'Invite' AS visitor_type_display,
                    ua.building_id,
                    ua.level_id,
                    ua.unit_id,
                    pb.building_name,
                    bl.level AS level_name,
                    bu.unit_no AS unit_name
                FROM invite_visitor iv
                LEFT JOIN user_access_details ua ON iv.user_id = ua.user_id
                LEFT JOIN property_building pb ON ua.building_id = pb.id
                LEFT JOIN building_level bl ON ua.level_id = bl.id
                LEFT JOIN building_units bu ON ua.unit_id = bu.id
            ) AS visitors
            WHERE 1=1
        """

        if id is not None:
            query += " AND id = :id"
            count_query += " AND id = :id"
            params["id"] = id

        if visitor_id is not None:
            query += " AND visitor_id = :visitor_id"
            count_query += " AND visitor_id = :visitor_id"
            params["visitor_id"] = visitor_id

        if visitor_type is not None:
            query += " AND visitor_type_raw = :visitor_type"
            count_query += " AND visitor_type_raw = :visitor_type"
            params["visitor_type"] = visitor_type

        if search:
            query += " AND (name LIKE :search OR visitor_id LIKE :search OR purpose_visit LIKE :search OR phone LIKE :search)"
            count_query += " AND (name LIKE :search OR visitor_id LIKE :search OR purpose_visit LIKE :search OR phone LIKE :search)"
            params["search"] = f"%{search}%"

        if building_id is not None:
            query += " AND building_id = :building_id"
            count_query += " AND building_id = :building_id"
            params["building_id"] = building_id

        if level_id is not None:
            query += " AND level_id = :level_id"
            count_query += " AND level_id = :level_id"
            params["level_id"] = level_id

        if unit_id is not None:
            query += " AND unit_id = :unit_id"
            count_query += " AND unit_id = :unit_id"
            params["unit_id"] = unit_id

        if start_date:
            query += " AND DATE(created_at) >= :start_date"
            count_query += " AND DATE(created_at) >= :start_date"
            params["start_date"] = start_date.date() if isinstance(start_date, datetime) else start_date

        if end_date:
            query += " AND DATE(created_at) <= :end_date"
            count_query += " AND DATE(created_at) <= :end_date"
            params["end_date"] = end_date.date() if isinstance(end_date, datetime) else end_date

        if is_qr_valid is not None:
            now_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
            if is_qr_valid:
                query += " AND JSON_UNQUOTE(JSON_EXTRACT(valid, '$.endTime')) >= :now_time"
                count_query += " AND JSON_UNQUOTE(JSON_EXTRACT(valid, '$.endTime')) >= :now_time"
            else:
                query += " AND JSON_UNQUOTE(JSON_EXTRACT(valid, '$.endTime')) < :now_time"
                count_query += " AND JSON_UNQUOTE(JSON_EXTRACT(valid, '$.endTime')) < :now_time"
            params["now_time"] = now_str

        if id is not None or visitor_id is not None:
            query += " ORDER BY created_at DESC LIMIT 1"
            result = db.execute(text(query), params).mappings().first()

            if result:
                row = dict(result)

                if row.get("valid") and isinstance(row["valid"], str):
                    try:
                        row["valid"] = json.loads(row["valid"])
                    except:
                        pass

                if "visitor_type_display" in row:
                    row["visitor_type"] = row.pop("visitor_type_display")
                if "visitor_type_raw" in row:
                    row.pop("visitor_type_raw")

                return [row], 1

            return [], 0

        total_count = db.execute(text(count_query), params).scalar() or 0

        if not download:
            query += " ORDER BY created_at DESC LIMIT :limit OFFSET :offset"
            params["limit"] = record_count
            params["offset"] = offset
        else:
            query += " ORDER BY created_at DESC"

        results = db.execute(text(query), params).mappings().all()

        data = []
        for row in results:
            row_dict = dict(row)

            if row_dict.get("valid") and isinstance(row_dict["valid"], str):
                try:
                    row_dict["valid"] = json.loads(row_dict["valid"])
                except:
                    pass

            if "visitor_type_display" in row_dict:
                row_dict["visitor_type"] = row_dict.pop("visitor_type_display")
            if "visitor_type_raw" in row_dict:
                row_dict.pop("visitor_type_raw")

            data.append(row_dict)

        return data, total_count

    except Exception as e:
        raise Exception(f"Error in list_visitors_service: {str(e)}")


def _fmt_dt(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def generate_visitors_excel_response(data: List[Dict[str, Any]], filename: str = "visitors.xlsx"):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "S.No",
            "Name",
            "Type",
            "Phone",
            "Purpose",
            "Block",
            "Level",
            "Unit",
        ]

        field_mapping = {
            "Name": "name",
            "Type": "visitor_type",
            "Phone": "phone",
            "Purpose": "purpose_visit",
            "Building": "building_name",
            "Level": "level_name",
            "Unit": "unit_name",
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for row_index, record in enumerate(data):
            row_num = row_index + 2
            for col_num, header in enumerate(headers, 1):
                if header == "S.No":
                    value = row_index + 1
                else:
                    field_name = field_mapping.get(header)
                    value = record.get(field_name, "") if field_name else ""
                    if header == "Created At":
                        value = _fmt_dt(value) or "-"
                    if value is None or value == "":
                        value = "-"

                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = cell_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(data) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")


def generate_visitors_pdf_response(data: List[Dict[str, Any]], filename: str = "visitors.pdf"):
    """Generate PDF response using the same UI style as activity logs."""
    try:
        pdf_buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        elements = []
        styles = getSampleStyleSheet()

        logo_url = get_logo_url_prod()
        try:
            logo = Image(logo_url)
            logo.drawHeight = 1.0 * inch
            logo.drawWidth = 2.0 * inch
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

        title_style = styles['Heading1']
        title_style.alignment = 1
        title_style.fontSize = 16
        title_style.spaceAfter = 0.2 * inch
        elements.append(Paragraph("Visitors Report", title_style))

        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            alignment=1,
            fontSize=10,
            textColor=colors.gray,
            spaceAfter=0.2 * inch
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.2 * inch))

        headers = [
            "S.No",
            "Name",
            "Type",
            "Phone",
            "Purpose",
            "Block",
            "Level",
            "Unit",
        ]

        page_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
        col_widths = [
            page_width * 0.06,
            page_width * 0.12,
            page_width * 0.12,
            page_width * 0.08,
            page_width * 0.10,
            page_width * 0.16,
            page_width * 0.14,
            page_width * 0.07,
            page_width * 0.07,
            page_width * 0.08,
        ]

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ])

        max_rows_per_page = 12
        total_records = len(data)

        for start in range(0, total_records, max_rows_per_page):
            end = start + max_rows_per_page
            chunk = data[start:end]

            table_data = [headers]
            for idx, record in enumerate(chunk, start=start + 1):
                row = [
                    str(idx),
                    record.get('name') or "-",
                    record.get('visitor_type') or "-",
                    record.get('phone') or "-",
                    record.get('purpose_visit') or "-",
                    record.get('building_name') or "-",
                    record.get('level_name') or "-",
                    record.get('unit_name') or "-",
                ]
                table_data.append(row)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(table_style)
            elements.append(table)

            if end < total_records:
                elements.append(PageBreak())

        elements.append(Spacer(1, 0.2 * inch))
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.darkblue
        )
        elements.append(Paragraph(f"Total Visitors: {len(data)}", summary_style))

        doc.build(elements)
        pdf_buffer.seek(0)

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")